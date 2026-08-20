"""
アフターコーディング支援ツール - デモ版
Streamlit Webアプリ
"""

import streamlit as st
import anthropic
import copy
import hashlib
import json
import re
import time
import random
import math
import io
import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.data_source import AxDataSource, StrRef
from datetime import datetime
from llm_client import call_llm, make_client, reset_token_usage, get_token_usage, get_last_error

APP_DIR = Path(__file__).parent

# コードブック策定・編集は判断の質が重要なためSonnet固定。
# コーディング（既存コードブックへの分類作業）はサイドバーで選択可能（精度・価格比較のため）。
CODEBOOK_MODEL = 'claude-sonnet-4-6'
CODING_MODEL   = 'claude-haiku-4-5'  # コーディングモデルのデフォルト（サイドバー未選択時など）
CODING_MODEL_OPTIONS = {
    'Haiku 4.5（安い・高速）':   'claude-haiku-4-5',
    'Sonnet 4.6（高精度・高コスト）': 'claude-sonnet-4-6',
}

CODING_STRICTNESS = 'standard'  # コーディング判定の厳密度デフォルト（サイドバー未選択時など）
CODING_STRICTNESS_OPTIONS = {
    '標準':                     'standard',
    '厳密（誤検出を避ける）':   'strict',
    '柔軟（見落としを避ける）': 'lenient',
}

# リスクチェック項目。チェックした項目だけをコーディング時にAIへ問い合わせる（未チェックはコスト0）。
# key: 内部フィールド名（LLMスキーマ・result内で使用） / label: UI表示名
# char: Excel「回答別コーディング結果」の一文字見出し / hint: チェックの説明
RISK_CHECK_OPTIONS = [
    {'key': 'claim',    'label': 'クレーム',   'char': 'ク', 'hint': '感情的で強いクレーム'},
    {'key': 'personal', 'label': '個人情報',   'char': '個', 'hint': '個人名・メールアドレス・電話番号・詳細住所'},
    {'key': 'org',      'label': '団体名',   'char': '団',
     'hint': '学校名・病院名・施設名としての具体的な固有名詞（例：八戸第3小学校、△△総合病院）が明記されている場合のみ該当。'
             '「校長と教頭と教務主導で」「ICT担当のパワハラ教員」「管理職３人」のように、役職・立場・職務内容を表す言葉だけで'
             '固有名詞が一切ない文は、絶対に該当としない。'},
    {'key': 'danger',   'label': '危険情報',   'char': '危', 'hint': '犯罪予告、自死予告、強い恨み、特定民族への差別'},
]

# 回答分類項目。センチメント・非該当と同じ「固定」区分（チェックボックスなし、常時判定）。
# llm=True: コーディング時にAIへ判定させる（llm_code_batchのスキーマ・プロンプトに常に含める）
# llm=False: 回答テキストから機械的に判定する（AIには問い合わせない）
ANSWER_TYPE_OPTIONS = [
    {'key': 'unclear',    'label': '不明',   'char': '不', 'llm': True,
     'hint': '特になし・とくになし・なし・わからない・答えたくない・意味の不明な文字入力'},
    {'key': 'unanswered', 'label': '無回答', 'char': '無', 'llm': False,
     'hint': '空欄'},
]

# 画面の縦棒グラフ（コード別GT集計）とExcelレポートで完全に同じ色をカテゴリごとに使うための
# 固定配色。st.plotly_chart()は既定でStreamlit自身のテーマ配色を自動適用するが、その並び順は
# ブラウザ／OSのライト・ダーク設定によって変わってしまう（同じデータでも色が入れ替わる）ため、
# ここで固定のcolor_discrete_mapとして明示的に指定し、画面表示・Excel・テーマ設定・再実行の
# あいだで常に同じ色になるようにする。
CHART_COLORS = ['0068C9', '83C9FF', 'FF2B2B', 'FFABAB', '29B09D',
                 '7DEFA1', 'FF8700', 'FFD16A', '6D3FC0', 'D5DAE5']


def _category_color_map(gt, codes):
    """カテゴリの出現件数合計が多い順に、CHART_COLORSを先頭から割り当てる。"""
    cat_total = {}
    for g in gt:
        cat_total[g['cat_id']] = cat_total.get(g['cat_id'], 0) + g['count']
    for c in codes:
        cat_total.setdefault(c['cat_id'], 0)
    cat_order = sorted(cat_total, key=lambda cid: -cat_total[cid])
    cat_color_full = {cid: CHART_COLORS[i % len(CHART_COLORS)] for i, cid in enumerate(cat_order)}
    return cat_order, cat_color_full


def _lighten_hex(hex_color, factor=0.65):
    """指定した割合(0〜1)だけ白に近づけた淡い色を返す（Excel出力・バブル図で共通利用）。"""
    r = int(hex_color[0:2], 16); g = int(hex_color[2:4], 16); b = int(hex_color[4:6], 16)
    r = round(r + (255 - r) * factor)
    g = round(g + (255 - g) * factor)
    b = round(b + (255 - b) * factor)
    return f'{r:02X}{g:02X}{b:02X}'


def _items_from_texts(texts):
    """1行1回答のプレーンテキスト入力を、Excel入力と同じitems形式に揃える（fa_no・attrsは空）。"""
    return [{'id': f'NO{i+1:03d}', 'text': t, 'fa_no': None, 'attrs': {}} for i, t in enumerate(texts)]


PROJECT_FILE_APP_TAG      = 'after_coder_project'
PROJECT_FILE_FORMAT_VER   = 1


def _build_project_file(result):
    """
    分析結果一式（RAWデータ・列マッピング・コードブック・コーディング結果を含むresult）を、
    後で「プロジェクトファイルを開く」から読み込んで再開できるJSON形式に変換する。
    pending_edit（未確定の編集案）は再開時に混乱を招くだけなので含めない。
    """
    payload_result = {k: v for k, v in result.items() if k != 'pending_edit'}
    return {
        'app':            PROJECT_FILE_APP_TAG,
        'format_version': PROJECT_FILE_FORMAT_VER,
        'saved_at':        datetime.now().isoformat(),
        'result':          payload_result,
    }


def _load_project_file(payload):
    """_build_project_fileで作られたJSONを読み込み、historyに積めるresult辞書として復元する。"""
    if not isinstance(payload, dict) or payload.get('app') != PROJECT_FILE_APP_TAG:
        raise ValueError('After Coderのプロジェクトファイルではないようです。')
    result = payload.get('result')
    if not isinstance(result, dict) or 'codebook' not in result or 'items' not in result:
        raise ValueError('プロジェクトファイルの内容が不完全です。')
    return result


st.set_page_config(
    page_title='アフターコーディング支援ツール',
    page_icon='📊',
    layout='wide'
)
st.logo(str(APP_DIR / 'after_coder_MARK.png'), size='large')

# ── スタイル ──────────────────────────────────────────────────────
st.markdown("""
<style>
.main-title { font-size: 28px; font-weight: bold; color: #2E5C8A; margin-bottom: 0; }
.sub-title  { font-size: 14px; color: #666; margin-bottom: 24px; }
.sidebar-title-wrap { margin-top: -10px; }
.sidebar-title-jp   { font-size: 15px !important; font-weight: bold; color: #FFFFFF; margin: 0; white-space: nowrap; }
.sidebar-title-en   { font-size: 40px !important; font-weight: bold; color: #72C6EF; margin: 2px 0 0 0; }
.step-label { font-size: 13px; font-weight: bold; color: #2E5C8A; }
.result-box { background: #f8f9fa; border-radius: 8px; padding: 16px; margin: 8px 0; }
div[data-testid="stMainBlockContainer"] { padding-top: 2rem !important; }
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════
# 認証
# ══════════════════════════════════════════════════

def authenticate(username, password):
    """
    ユーザーID・パスワードはコードに直接書かず、st.secretsから読む。
    ローカルは.streamlit/secrets.toml（Git管理外）、本番はStreamlit Community Cloudの
    アプリ設定「Secrets」に、以下の形式で設定する。
        admin_username = "admin"
        admin_password = "..."
        [users]
        someuser = "..."
    secrets.toml自体が無い場合や該当キーが無い場合はNone/空辞書扱いとなり、認証は常に失敗する
    （＝資格情報を設定しない限りログインできない、フェイルセーフな挙動）。
    """
    admin_username = st.secrets.get('admin_username', 'admin')
    admin_password = st.secrets.get('admin_password')
    if username == admin_username:
        return admin_password is not None and password == admin_password
    users = st.secrets.get('users', {})
    return users.get(username) == password


st.session_state.setdefault('authenticated', False)
st.session_state.setdefault('username', None)
st.session_state.setdefault('history', [])
st.session_state.setdefault('history_counter', 0)
st.session_state.setdefault('active_history_id', None)
st.session_state.setdefault('texts_count', 0)
st.session_state.setdefault('xlsx_items', [])
st.session_state.setdefault('coding_job', None)
st.session_state.setdefault('diagnostic_job', None)


def render_login():
    st.markdown('<p class="main-title">👻 アフターコーディング支援ツール</p>', unsafe_allow_html=True)
    st.markdown('<p class="sub-title">ログインしてください</p>', unsafe_allow_html=True)
    _, mid, _ = st.columns([1, 1, 1])
    with mid:
        with st.form('login_form'):
            username = st.text_input('ユーザーID')
            password = st.text_input('パスワード', type='password')
            submitted = st.form_submit_button('ログイン', type='primary', width='stretch')
            if submitted:
                if authenticate(username, password):
                    st.session_state.authenticated = True
                    st.session_state.username = username
                    st.rerun()
                else:
                    st.error('ユーザーIDまたはパスワードが正しくありません')


if not st.session_state.authenticated:
    render_login()
    st.stop()


# ══════════════════════════════════════════════════
# LLM処理関数
# ══════════════════════════════════════════════════

def llm_generate_codebook(client, items, max_codes, q_name, data_context=''):
    text = '\n'.join(f'{x["id"]}: {x["text"]}' for x in items)
    context_str = f'\n【調査の背景・特徴】\n{data_context}\n' if data_context.strip() else ''
    prompt = f"""「{q_name}」の自由回答（{len(items)}件）からコードブックを作成してください。
{context_str}
【ルール】
- 中間カテゴリ7個前後（最大10個）
- カテゴリ内コード最大10個、コード総数は必ず{max_codes}個以内に収める（絶対に超えないこと。類似する主題は1つのコードに統合する）
- 1コード＝1主題
- コードID形式: CAT01.../C0101...（コードIDの接頭辞は英字1文字の「C」のみ。「CO」のように2文字以上にしない）

【回答リスト】
{text}"""

    schema = {
        'type': 'object',
        'properties': {
            'categories': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'cat_id':   {'type': 'string'},
                        'cat_name': {'type': 'string'},
                        'codes': {
                            'type': 'array',
                            'items': {
                                'type': 'object',
                                'properties': {
                                    'code_id':    {'type': 'string'},
                                    'code_name':  {'type': 'string'},
                                    'definition': {'type': 'string'},
                                },
                                'required': ['code_id', 'code_name', 'definition'],
                            }
                        }
                    },
                    'required': ['cat_id', 'cat_name', 'codes'],
                }
            }
        },
        'required': ['categories'],
    }
    return call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)


def llm_generate_codebook_topdown(client, data_context, q_name, max_codes):
    """（方式B用）実データを見る前に、調査の背景・特徴だけからカテゴリ骨格を生成"""
    prompt = f"""「{q_name}」の自由回答について、実際の回答データを見る前に、
調査の背景・特徴だけから想定されるコードブックの骨格（カテゴリと想定コード名）を作成してください。

【調査の背景・特徴】
{data_context}

【ルール】
- 中間カテゴリ7個前後（最大10個）
- カテゴリごとに想定コード名を列挙（1カテゴリあたり最大10個、コード総数は必ず{max_codes}個以内に収める。絶対に超えないこと）
- コードIDは不要、コード名のみ列挙
- カテゴリIDはCAT01形式"""

    schema = {
        'type': 'object',
        'properties': {
            'categories': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'cat_id':         {'type': 'string'},
                        'cat_name':       {'type': 'string'},
                        'expected_codes': {'type': 'array', 'items': {'type': 'string'}},
                    },
                    'required': ['cat_id', 'cat_name', 'expected_codes'],
                }
            }
        },
        'required': ['categories'],
    }
    return call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)


def llm_elaborate_skeleton(client, skeleton, sample_items, max_codes, q_name, data_context=''):
    """（方式B用）トップダウン骨格を実データサンプルで具体化し、定義を付与したコードブックを確定"""
    skeleton_text = '\n'.join(
        f'{cat.get("cat_id", "")}: {cat.get("cat_name", "")} → ' + '、'.join(cat.get('expected_codes', []) or [])
        for cat in skeleton.get('categories', [])
    )
    sample_text = '\n'.join(f'{x["id"]}: {x["text"]}' for x in sample_items)
    context_str = f'\n【調査の背景・特徴】\n{data_context}\n' if data_context.strip() else ''
    prompt = f"""「{q_name}」のコードブックを確定してください。
以下は調査の背景から事前に想定したカテゴリ骨格です。骨格のカテゴリ構成（cat_id・cat_name）は維持しながら、
実際の回答サンプルを踏まえて各コードに定義を付与し、必要に応じてコードを追加・調整してください。

【カテゴリ骨格（トップダウンで抽出した想定コード）】
{skeleton_text}
{context_str}
【実際の回答サンプル（{len(sample_items)}件）】
{sample_text}

【ルール】
- 骨格のカテゴリ（cat_id・cat_name）は維持する
- 想定コードに定義を付与する（サンプルに現れない想定コードも残してよい）
- サンプルに現れる主題で骨格にないものは追加する（コード総数は必ず{max_codes}個以内に収める。絶対に超えないこと。追加すると上限を超える場合は既存の想定コードと統合する）
- 1コード＝1主題
- コードID形式: CAT01.../C0101...（コードIDの接頭辞は英字1文字の「C」のみ。「CO」のように2文字以上にしない）"""

    schema = {
        'type': 'object',
        'properties': {
            'categories': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'cat_id':   {'type': 'string'},
                        'cat_name': {'type': 'string'},
                        'codes': {
                            'type': 'array',
                            'items': {
                                'type': 'object',
                                'properties': {
                                    'code_id':    {'type': 'string'},
                                    'code_name':  {'type': 'string'},
                                    'definition': {'type': 'string'},
                                },
                                'required': ['code_id', 'code_name', 'definition'],
                            }
                        }
                    },
                    'required': ['cat_id', 'cat_name', 'codes'],
                }
            }
        },
        'required': ['categories'],
    }
    return call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)


def llm_extract_topics(client, items, q_name, data_context=''):
    """（方式C用）回答バッチから主題（トピック）を網羅的に抽出"""
    text = '\n'.join(f'{x["id"]}: {x["text"]}' for x in items)
    context_str = f'\n【調査の背景・特徴】\n{data_context}\n' if data_context.strip() else ''
    prompt = f"""「{q_name}」の回答（{len(items)}件）に含まれる主題（トピック）をすべて抽出してください。
{context_str}
【ルール】
- 回答ごとに1つ以上の主題を短いフレーズで抽出する
- 似た内容は同じ表現にまとめてよい
- 網羅性を優先し、細かい主題も漏らさず列挙する

【回答リスト】
{text}"""

    schema = {
        'type': 'object',
        'properties': {
            'topics': {'type': 'array', 'items': {'type': 'string'}}
        },
        'required': ['topics'],
    }
    result = call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)
    if result and isinstance(result, dict):
        return result.get('topics', [])
    return []


def llm_reduce_topics(client, topics, q_name, data_context=''):
    """（方式C用）主題リストのチャンクを、同義・類似表現をまとめて簡潔なユニークリストに整理する"""
    topics_text = '\n'.join(f'- {t}' for t in topics)
    context_str = f'\n【調査の背景・特徴】\n{data_context}\n' if data_context.strip() else ''
    prompt = f"""「{q_name}」の回答から抽出された主題リスト（{len(topics)}件）を整理してください。
{context_str}
【主題リスト】
{topics_text}

【ルール】
- 同義・類似する主題はひとつにまとめる
- 表現はできるだけ簡潔な短いフレーズにする
- 異なる主題の情報を失わないよう、まとめすぎない"""

    schema = {
        'type': 'object',
        'properties': {
            'topics': {'type': 'array', 'items': {'type': 'string'}}
        },
        'required': ['topics'],
    }
    result = call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)
    if result and isinstance(result, dict):
        return result.get('topics', [])
    return topics  # 失敗時は元のチャンクをそのまま返し情報を失わない


def llm_consolidate_topics(client, topics, max_codes, q_name, data_context=''):
    """（方式C用）全件から蓄積した主題リストを統合し、定義・キーワード付きのコードブックを確定"""
    topics_text = '\n'.join(f'- {t}' for t in topics)
    context_str = f'\n【調査の背景・特徴】\n{data_context}\n' if data_context.strip() else ''
    prompt = f"""「{q_name}」の全回答から抽出された主題リスト（{len(topics)}件、類似表現含む）を統合し、
最終的なコードブックを確定してください。
{context_str}
【主題リスト】
{topics_text}

【ルール】
- 類似・重複する主題を統合し、中間カテゴリ7個前後（最大10個）に整理する
- カテゴリ内コード最大10個、コード総数は必ず{max_codes}個以内に収める（絶対に超えないこと。主題リストの件数が多い場合は、内容が近い主題同士を積極的に1つのコードにまとめて件数を減らす）
- 1コード＝1主題
- 各コードに定義と、判定に役立つキーワードを2〜5個付与する
- コードID形式: CAT01.../C0101...（コードIDの接頭辞は英字1文字の「C」のみ。「CO」のように2文字以上にしない）"""

    schema = {
        'type': 'object',
        'properties': {
            'categories': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'cat_id':   {'type': 'string'},
                        'cat_name': {'type': 'string'},
                        'codes': {
                            'type': 'array',
                            'items': {
                                'type': 'object',
                                'properties': {
                                    'code_id':    {'type': 'string'},
                                    'code_name':  {'type': 'string'},
                                    'definition': {'type': 'string'},
                                    'keywords':   {'type': 'array', 'items': {'type': 'string'}},
                                },
                                'required': ['code_id', 'code_name', 'definition'],
                            }
                        }
                    },
                    'required': ['cat_id', 'cat_name', 'codes'],
                }
            }
        },
        'required': ['categories'],
    }
    return call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)


def llm_detect_new_codes(client, items, codebook, max_codes, q_name):
    text = '\n'.join(f'{x["id"]}: {x["text"]}' for x in items)
    existing = '\n'.join(
        f'{c["code_id"]}({cat["cat_id"]}:{cat["cat_name"]}):{c["code_name"]}'
        for cat in codebook.get('categories', [])
        for c in cat.get('codes', [])
    )
    cat_list = '\n'.join(
        f'{cat["cat_id"]}: {cat["cat_name"]}'
        for cat in codebook.get('categories', [])
    )
    cur = sum(len(c['codes']) for c in codebook.get('categories', []))
    prompt = f"""「{q_name}」の回答（{len(items)}件）に既存コードで対応できない新主題がありますか？

【既存カテゴリ一覧】
{cat_list}

【既存コード（{cur}個、残り{max_codes-cur}個）】
{existing}

【回答】
{text}

新主題がある場合のみ返してください。なければnew_codesを空配列に。
cat_idは必ず既存カテゴリ一覧のCAT01形式を使うこと。
code_idは既存コードと同じC0101形式で採番すること（接頭辞は英字1文字の「C」のみ。「CO」のように2文字以上にしない）。"""

    schema = {
        'type': 'object',
        'properties': {
            'new_codes': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'cat_id':    {'type': 'string'},
                        'code_id':   {'type': 'string'},
                        'code_name': {'type': 'string'},
                        'definition':{'type': 'string'},
                    },
                    'required': ['cat_id', 'code_id', 'code_name', 'definition'],
                }
            }
        },
        'required': ['new_codes'],
    }
    result = call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)
    if result is None:
        return []
    return result.get('new_codes', [])


_SENTENCE_SPLIT_RE = re.compile(r'(?<=[。！？])')


def _split_sentences(text):
    """
    テキストを文単位に分割する（句点・感嘆符・疑問符の直後で区切るルールベースの文境界検出。
    形態素解析＝単語分割ではなく、文の切れ目を見つけるだけなので外部ライブラリは使わない）。
    末尾に区切り記号が無い断片も1文として扱う。空白のみ・空文字列の断片は除外する。
    """
    if not text:
        return []
    parts = _SENTENCE_SPLIT_RE.split(text)
    return [p.strip() for p in parts if p.strip()]


def llm_code_batch(client, items, codes, q_name, model=CODING_MODEL, enabled_risks=None,
                    strictness=CODING_STRICTNESS):
    """
    コーディング本体。同じコードブック（system側）を1回のコーディング実行中に何十回も
    使い回すため、コードブック・ルールをsystemに分離しcache_control（プロンプトキャッシュ）を
    効かせ、2回目以降のバッチでコードブック分のコストを大幅に削減する。
    modelはサイドバーの「コーディングモデル」選択に従う（デフォルトはHaiku=CODING_MODEL）。
    精度・価格を比較したい場合は、同じコードブックのまま方式を変えて2回実行し、
    作業履歴で比較する使い方を想定している。
    enabled_risksはサイドバーの「リスクチェック」でチェックされた項目のkeyのリスト。
    未チェックの項目はプロンプト・スキーマに一切含めない（AIへの問い合わせ自体をしない＝コスト増なし）。
    strictnessはサイドバーの「コーディング判定の厳密度」（standard/strict/lenient）。
    コード付与の判定基準にのみ効かせ、リスクチェック・回答分類の判定基準は変えない。
    """
    enabled_risks = enabled_risks or []
    risk_opts = [o for o in RISK_CHECK_OPTIONS if o['key'] in enabled_risks]
    # 回答分類（不明・無回答）のうちAI判定が必要なもの（無回答は回答テキストの空欄チェックで
    # 機械的に判定するため、_code_items側で処理しここには含めない）
    llm_answer_types = [o for o in ANSWER_TYPE_OPTIONS if o.get('llm')]

    # 定義文は全文をそのまま渡す（以前は[:25]で先頭25文字に切り詰めていたが、区別のポイントが
    # それより後ろに書かれている定義が多く、似たコード同士の混同の一因になっていた。コードブックは
    # systemプロンプト側でキャッシュされるため、全文にしてもコスト増は初回バッチ分のみで軽微）。
    code_list = '\n'.join(
        f'{c["code_id"]}（{c["cat_name"]}）: {c["code_name"]} / {c["definition"]}'
        for c in codes
    )
    answer_type_rule = ''
    if llm_answer_types:
        lines = '\n'.join(f"- {o['label']}: {o['hint']}" for o in llm_answer_types)
        answer_type_rule = f"""

【回答分類】次の該当有無も回答ごとに判定してください（該当すればtrue、しなければfalse）
{lines}"""

    risk_rule = ''
    if risk_opts:
        risk_lines = '\n'.join(f"- {o['label']}: {o['hint']}" for o in risk_opts)
        risk_rule = f"""

【リスクチェック】次の該当有無も回答ごとに判定してください（該当すればtrue、しなければfalse）
{risk_lines}"""

    strictness_rule = ''
    if strictness == 'strict':
        strictness_rule = """

【コード判定基準】コードの定義・キーワードに明確に一致する場合のみ付与してください。文脈からの推測だけでは付与しないでください。"""
    elif strictness == 'lenient':
        strictness_rule = """

【コード判定基準】コードのテーマに関連する内容であれば、表現が異なっていても幅広く付与してください。文脈から示唆される内容も対象に含めてください。"""

    system_prompt = f"""「{q_name}」の回答にコードブックに基づいてコーディングしてください。
回答は句点等で区切った文ごとに番号を振って渡します。回答全体の文脈を踏まえた上で判断してください
（1文だけを単独で読むのではなく、前後の文もヒントにしてください）。ただし、コード・センチメントの
判定結果は文ごとに返してください。

【コードブック】
{code_list}

【ルール】
- 1文に複数コード付与可。該当なしはcodesを空配列
- sentimentは文ごとに positive/negative/neutral のいずれか
- key_sentence_idxには、その回答の中で回答者が最も伝えたいこと・結論と言える文の番号を1つ指定する
  （該当する文がない、または回答が空の場合は0）{answer_type_rule}{risk_rule}{strictness_rule}"""

    sentences_by_id = {x['id']: _split_sentences(x['text']) for x in items}

    def _numbered(sentences):
        if not sentences:
            return '(本文なし)'
        return '\n'.join(f'[{i+1}] {s}' for i, s in enumerate(sentences))

    items_text = '\n\n'.join(f'{x["id"]}:\n{_numbered(sentences_by_id[x["id"]])}' for x in items)
    prompt = f"""【回答】
{items_text}"""

    sentence_properties = {
        'idx':       {'type': 'integer'},
        'sentiment': {'type': 'string', 'enum': ['positive', 'negative', 'neutral']},
        'codes':     {'type': 'array', 'items': {'type': 'string'}},
    }
    result_properties = {
        'id': {'type': 'string'},
        'sentences': {
            'type': 'array',
            'items': {
                'type': 'object',
                'properties': sentence_properties,
                'required': ['idx', 'sentiment', 'codes'],
            },
        },
        'key_sentence_idx': {'type': 'integer'},
    }
    for o in llm_answer_types:
        result_properties[o['key']] = {'type': 'boolean'}
    for o in risk_opts:
        result_properties[o['key']] = {'type': 'boolean'}

    schema = {
        'type': 'object',
        'properties': {
            'results': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': result_properties,
                    'required': ['id', 'sentences', 'key_sentence_idx'] + [o['key'] for o in llm_answer_types],
                }
            }
        },
        'required': ['results'],
    }
    result = call_llm(client, prompt, schema, 'Anthropic', model,
                       system=system_prompt, cache_system=True)
    raw_results = result.get('results', []) if result and isinstance(result, dict) else []

    # AIが返す各結果のidは信頼せず、返ってきた件数が送った件数と一致する場合は、items（送った
    # 順序）に基づいて強制的に付け直す。文番号付きの複数行プロンプト（[1] 文...\n[2] 文...）に
    # 変更して以降、実データで「AIが返すidが元のidと一致せず、無回答判定・自由回答一覧・文単位
    # データの照合が総崩れになる」不具合が確認された。AIは文字列（英数字ID）を1文字も違わず
    # 書き写すことより、渡された順序を保つことの方が信頼性が高いという前提に基づく対策
    # （2026-08-16、ユーザー提供の実データで発覚・修正）。件数が一致しない場合（AIが一部の
    # 回答を欠落・重複させた場合）は、この付け替えができないため、AIの返したid文字列に頼る
    # 従来の挙動にフォールバックする。
    if len(raw_results) == len(items):
        for item, r in zip(items, raw_results):
            r['id'] = item['id']

    # 文ごとの判定（sentences）を回答単位のcodes/sentimentに集約し、既存の呼び出し元
    # （aggregate_results・Excel生成など）が今まで通り動くようにする。集約ルール：
    # codesは全文の和集合、sentimentは全文が同じ値ならその値、割れていれば'mixed'（混在）。
    # 文テキスト・結論文（key_sentence）は文番号(idx)から元の文リストを引いて埋め込む
    # （AIには文番号だけを返させ、文そのものを書き写させない＝原文と完全一致する引用になる）。
    for r in raw_results:
        sentences = sentences_by_id.get(r.get('id'), [])
        sent_list = r.get('sentences', [])
        for s in sent_list:
            idx = s.get('idx')
            s['text'] = sentences[idx - 1] if isinstance(idx, int) and 1 <= idx <= len(sentences) else ''

        all_codes = []
        for s in sent_list:
            for cid in s.get('codes', []):
                if cid not in all_codes:
                    all_codes.append(cid)
        r['codes'] = all_codes

        sentiments = {s.get('sentiment') for s in sent_list if s.get('sentiment')}
        if len(sentiments) == 1:
            r['sentiment'] = next(iter(sentiments))
        elif len(sentiments) > 1:
            r['sentiment'] = 'mixed'
        else:
            r['sentiment'] = 'neutral'

        key_idx = r.get('key_sentence_idx')
        r['key_sentence'] = sentences[key_idx - 1] if isinstance(key_idx, int) and 1 <= key_idx <= len(sentences) else ''

    return raw_results


def llm_edit_codebook(client, codebook, instruction, q_name):
    """
    コードブック編集機能：チャット指示に基づき、統合・改名・再定義・追加・削除などの編集を行う。
    生データは参照しない（新規コードの発見は行わず、既存コードブックの整理のみ）。
    """
    codebook_text = json.dumps(codebook, ensure_ascii=False)
    prompt = f"""「{q_name}」のコードブックを、次の指示に従って編集してください。

【現在のコードブック】
{codebook_text}

【編集指示】
{instruction}

【ルール】
- 指示された編集（統合・改名・再定義・コードの追加・削除など）のみを行う
- 指示にない部分はそのまま維持する
- 生データは参照できないため、指示にない新規コードの発見は行わない
- コードID・カテゴリIDの形式は維持する（新規追加時はCAT.../C0101形式で採番する。コードIDの接頭辞は英字1文字の「C」のみ。「CO」のように2文字以上にしない）
- 編集後のコードブック全体を返す"""

    schema = {
        'type': 'object',
        'properties': {
            'categories': {
                'type': 'array',
                'items': {
                    'type': 'object',
                    'properties': {
                        'cat_id':   {'type': 'string'},
                        'cat_name': {'type': 'string'},
                        'codes': {
                            'type': 'array',
                            'items': {
                                'type': 'object',
                                'properties': {
                                    'code_id':    {'type': 'string'},
                                    'code_name':  {'type': 'string'},
                                    'definition': {'type': 'string'},
                                    'keywords':   {'type': 'array', 'items': {'type': 'string'}},
                                },
                                'required': ['code_id', 'code_name', 'definition'],
                            }
                        }
                    },
                    'required': ['cat_id', 'cat_name', 'codes'],
                }
            }
        },
        'required': ['categories'],
    }
    return call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)


_PROPOSED_CODE_SCHEMA_PROPS = {
    'code_name':  {'type': 'string'},
    'definition': {'type': 'string'},
    'keywords':   {'type': 'array', 'items': {'type': 'string'}},
}


def llm_propose_merge(client, code_a, code_b, q_name):
    """
    統合候補の2コードの内容だけをもとに、統合後の1コード（名称・定義・キーワード）を提案する。
    llm_edit_codebookと違い、コードブック全体は編集せず、統合後の1コードの中身のみを返す
    （2026-08-19、コードブック編集タブ③の仕様変更：コード単位でユーザーが内容を確認・編集
    してから個別に採用できるようにするため）。
    """
    prompt = f"""「{q_name}」の自由回答コーディングで使っている、以下の2つのコードを1つに統合してください。

【コード1】
名称：{code_a['code_name']}
定義：{code_a.get('definition', '')}

【コード2】
名称：{code_b['code_name']}
定義：{code_b.get('definition', '')}

【ルール】
- 両方の内容を包含する、統合後のコード名・定義を1つ作成する
- 定義は元の2つの定義の要点を踏まえ、簡潔かつ具体的にする"""

    schema = {
        'type': 'object',
        'properties': _PROPOSED_CODE_SCHEMA_PROPS,
        'required': ['code_name', 'definition'],
    }
    return call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)


def llm_propose_split(client, code, q_name):
    """
    分割候補の1コードの内容だけをもとに、分割後の2コード（名称・定義・キーワード）を提案する。
    llm_propose_mergeと対になる関数（2026-08-19追加）。
    """
    prompt = f"""「{q_name}」の自由回答コーディングで使っている、以下のコードは範囲が広すぎるため、
内容に応じて2つの具体的なコードに分割してください。

【分割対象のコード】
名称：{code['code_name']}
定義：{code.get('definition', '')}

【ルール】
- 元のコードがカバーしていた内容を、過不足なく2つに分ける
- それぞれに適切なコード名・定義を付ける"""

    schema = {
        'type': 'object',
        'properties': {
            'codes': {
                'type': 'array',
                'items': {'type': 'object', 'properties': _PROPOSED_CODE_SCHEMA_PROPS,
                          'required': ['code_name', 'definition']},
                'minItems': 2, 'maxItems': 2,
            }
        },
        'required': ['codes'],
    }
    return call_llm(client, prompt, schema, 'Anthropic', CODEBOOK_MODEL)


def _next_code_id(codebook, cat_id, extra_taken=()):
    """指定カテゴリ内で使われていない次のコードID（C0101形式）を採番する。
    同じ採用操作で複数の新コードを同時に作る場合（分割の2件など）は、直前に採番した
    コードIDの番号部分をextra_takenに渡すことで衝突を避ける。"""
    cat_num = cat_id[3:] if cat_id and cat_id.startswith('CAT') else '01'
    nums = [int(x) for x in extra_taken]
    for cat in codebook.get('categories', []):
        if cat.get('cat_id') == cat_id:
            for c in cat.get('codes', []):
                cid = c.get('code_id', '')
                if len(cid) >= 5 and cid[1:3] == cat_num:
                    try:
                        nums.append(int(cid[3:5]))
                    except ValueError:
                        pass
    return f'C{cat_num}{max(nums, default=0) + 1:02d}'


def _apply_staged_changes(result, staged_changes):
    """
    「📝 コードブックを書き換える」ボタンの処理。統合・分割候補の「採用する」は、この関数が
    呼ばれるまでコードブックを一切書き換えない（2026-08-19、仕様変更）。
    以前は候補ごとに採用と同時にコードブックへ即時反映していたが、そのたびにコードブック
    バージョン（edit_logの件数）が進み、他の未処理の候補が「候補作成後にコードブックが
    変わった」と判定されて軒並み使用不可になってしまう不具合があった（ユーザー報告：
    「統合作業を行ったあとに分割作業を行うと使用不可になり、作業が進まない」）。
    そこで「採用する」は変更内容をstaged_changes（result['pending_codebook_changes']）に
    積むだけにし、実際の書き換えはこの関数を呼ぶ「コードブックを書き換える」ボタンを
    押した時点でまとめて1回だけ行うようにした。これにより、書き換えるまでは
    codebook_versionが進まず、複数の統合・分割を続けて採用してから最後にまとめて
    反映できる。

    staged_changesは1件ずつ順番に適用する（同じカテゴリに複数件追加される場合でも、
    _next_code_idを直前までの適用結果に対して計算し直すため、コードIDが衝突しない）。
    適用前に、新しく追加されるコード名どうし、および削除されない既存コードとの重複が
    無いかを検証する（ユーザー要望：「コード名の重複などはこの段階でチェックしたら」）。
    重複がある場合は何も変更せず、(None, 重複名のリスト) を返す。
    成功時は (新しいcodebook, []) を返す。
    """
    new_codebook = copy.deepcopy(result['codebook'])

    all_remove_ids = set()
    for chg in staged_changes:
        all_remove_ids.update(chg['remove_ids'])
    existing_names = {
        c['code_name'] for cat in new_codebook.get('categories', []) for c in cat.get('codes', [])
        if c['code_id'] not in all_remove_ids
    }
    seen, dupes = set(), set()
    for chg in staged_changes:
        for nc in chg['new_codes']:
            name = nc['code_name']
            if name in existing_names or name in seen:
                dupes.add(name)
            seen.add(name)
    if dupes:
        return None, sorted(dupes)

    edit_log = result.setdefault('edit_log', [{'instruction': '（初期状態）', 'codebook': result['codebook']}])
    for chg in staged_changes:
        for cat in new_codebook.get('categories', []):
            cat['codes'] = [c for c in cat.get('codes', []) if c['code_id'] not in chg['remove_ids']]
        taken, add_codes = [], []
        for nc in chg['new_codes']:
            cid = _next_code_id(new_codebook, chg['target_cat_id'], extra_taken=taken)
            taken.append(cid[3:5])
            add_codes.append({**nc, 'code_id': cid})
        for cat in new_codebook.get('categories', []):
            if cat.get('cat_id') == chg['target_cat_id']:
                cat['codes'].extend(add_codes)
                break
        edit_log.append({'instruction': chg['label'], 'codebook': copy.deepcopy(new_codebook)})

    result['codebook'] = new_codebook
    result['codes'] = [
        {**c, 'cat_id': cat['cat_id'], 'cat_name': cat['cat_name']}
        for cat in new_codebook.get('categories', [])
        for c in cat.get('codes', [])
    ]
    return new_codebook, []


# ══════════════════════════════════════════════════
# 集計・Excel出力関数
# ══════════════════════════════════════════════════

def aggregate_results(codes, results, total, risk_keys=None):
    """
    コード別GT集計とセンチメント集計・非該当（どのコードも付与されなかった回答）件数、
    回答分類（不明・無回答）とリスクチェック（risk_keysで有効な項目のみ）の該当件数を算出
    """
    risk_keys   = risk_keys or []
    code_counts = {c['code_id']: 0 for c in codes}
    sent_counts = {'positive': 0, 'negative': 0, 'neutral': 0, 'mixed': 0}
    risk_counts = {k: 0 for k in risk_keys}
    answer_type_counts = {o['key']: 0 for o in ANSWER_TYPE_OPTIONS}
    result_map  = {r['id']: r for r in results}
    unassigned  = 0

    for rid, res in result_map.items():
        sent = res.get('sentiment', 'neutral')
        if sent in sent_counts:
            sent_counts[sent] += 1
        res_codes = res.get('codes', [])
        if not res_codes:
            unassigned += 1
        for cid in res_codes:
            if cid in code_counts:
                code_counts[cid] += 1
        for k in risk_keys:
            if res.get(k):
                risk_counts[k] += 1
        for o in ANSWER_TYPE_OPTIONS:
            if res.get(o['key']):
                answer_type_counts[o['key']] += 1

    gt = []
    for cat in codes:
        cnt = code_counts.get(cat['code_id'], 0)
        pct = cnt / total * 100 if total > 0 else 0
        gt.append({
            'cat_id':    cat['cat_id'],
            'cat_name':  cat['cat_name'],
            'code_id':   cat['code_id'],
            'code_name': cat['code_name'],
            'count':     cnt,
            'pct':       round(pct, 1),
            'definition':cat.get('definition', ''),
        })
    gt.sort(key=lambda x: x['count'], reverse=True)
    return gt, sent_counts, unassigned, risk_counts, answer_type_counts


CODEBOOK_CSV_COLUMNS = ['カテゴリID', 'カテゴリ名', 'コードID', 'コード名', '定義', 'キーワード']


def _codebook_rows(codebook, gt_by_code=None, include_stats=False):
    """コードブックをカテゴリ→コードの行データ（dictのリスト）に変換する共通処理"""
    gt_by_code = gt_by_code or {}
    rows = []
    for cat in codebook.get('categories', []):
        for c in cat.get('codes', []):
            keywords = c.get('keywords', [])
            if isinstance(keywords, list):
                keywords = '; '.join(keywords)
            row = {
                'カテゴリID': cat.get('cat_id', ''),
                'カテゴリ名': cat.get('cat_name', ''),
                'コードID':   c.get('code_id', ''),
                'コード名':   c.get('code_name', ''),
                '定義':       c.get('definition', ''),
                'キーワード': keywords or '',
            }
            if include_stats:
                stat = gt_by_code.get(c.get('code_id'), {})
                row = {'件数': stat.get('count', 0), '出現率(%)': stat.get('pct', 0.0), **row}
            rows.append(row)
    return rows


def _diff_codebook(old_codebook, new_codebook):
    """
    2つのコードブックをコードID単位で比較し、削除・追加・変更（コード名または定義が違う）された
    コードをそれぞれ返す（削除リスト, 追加リスト, 変更リスト[(旧,新), ...]）。
    編集案プレビューで「具体的に何が変わったか」を示すために使う（コードブック全体を毎回丸ごと
    表示するだけでは、コード数が多いと変更箇所が埋もれて分からないため）。
    """
    def flatten(cb):
        out = {}
        for cat in cb.get('categories', []):
            for c in cat.get('codes', []):
                out[c.get('code_id')] = {**c, 'cat_name': cat.get('cat_name', '')}
        return out

    old_codes = flatten(old_codebook)
    new_codes = flatten(new_codebook)

    removed = [old_codes[cid] for cid in old_codes if cid not in new_codes]
    added   = [new_codes[cid] for cid in new_codes if cid not in old_codes]
    changed = [
        (old_codes[cid], new_codes[cid])
        for cid in old_codes
        if cid in new_codes and (
            old_codes[cid].get('code_name') != new_codes[cid].get('code_name')
            or old_codes[cid].get('definition') != new_codes[cid].get('definition')
        )
    ]
    return removed, added, changed


def render_codebook_structure(codebook, gt_by_code=None, key=None):
    """
    コードブックの構造（件数・出現率(%)・カテゴリID・カテゴリ名・コードID・コード名・定義・キーワード）を、
    折りたたまず常時表示する。件数・出現率はgt_by_codeを渡せばコーディング結果を反映し、
    渡さない（または未コーディングの）コードは0として表示する。編集直後もその場で最新の内容が確認できる。
    セルをダブルクリックするとテキストを全文選択・コピーできる（st.data_editorを表示専用に使用。
    ここでの編集内容は保存されない＝実際のコードブックには反映されない）。
    表右上のツールバーからCSVダウンロードでき、そのCSVは「既存のコードブックを使用」で再読み込みできる
    （件数・出現率の列は読み込み時に無視される）。
    """
    import pandas as pd
    rows = _codebook_rows(codebook, gt_by_code, include_stats=True)
    n_cats = len(codebook.get('categories', []))
    st.caption(f'カテゴリ{n_cats}／コード{len(rows)}　※セルをダブルクリックするとテキストをコピーできます（ここでの編集内容は保存されません）')
    st.data_editor(pd.DataFrame(rows), width='stretch', hide_index=True, key=key)


def parse_codebook_csv(file_bytes):
    """
    render_codebook_structureの表からダウンロードしたCSV
    （カテゴリID,カテゴリ名,コードID,コード名,定義,キーワード。件数・出現率列があっても無視する）を
    コードブック構造（{'categories': [...]}）に復元する
    """
    import pandas as pd
    for enc in ('utf-8-sig', 'cp932'):
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), encoding=enc, dtype=str, keep_default_na=False)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError('文字コードを判定できませんでした（UTF-8またはShift-JISで保存し直してください）')

    missing = [c for c in CODEBOOK_CSV_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f'必要な列が見つかりません: {", ".join(missing)}')

    categories = {}
    for _, row in df.iterrows():
        cat_id = row['カテゴリID'].strip()
        if not cat_id:
            continue
        if cat_id not in categories:
            categories[cat_id] = {'cat_id': cat_id, 'cat_name': row['カテゴリ名'].strip(), 'codes': []}
        keywords = [k.strip() for k in row['キーワード'].split(';') if k.strip()]
        categories[cat_id]['codes'].append({
            'code_id':    row['コードID'].strip(),
            'code_name':  row['コード名'].strip(),
            'definition': row['定義'].strip(),
            'keywords':   keywords,
        })
    return {'categories': list(categories.values())}


def create_excel(q_name, gt, sent_counts, total, results, items, codes, unassigned=0,
                  risk_counts=None, enabled_risks=None, answer_type_counts=None):
    """
    ローカル版仮集計シートと同じレイアウトでExcelを生成。
    リスクチェック（risk_counts/enabled_risks）・回答分類（answer_type_counts、不明・無回答）は
    「特記情報集計」への項目追加と、「回答別コーディング結果」の一文字フラグ列に反映する。
    """
    risk_counts        = risk_counts or {}
    enabled_risks      = enabled_risks or []
    risk_opts          = [o for o in RISK_CHECK_OPTIONS if o['key'] in enabled_risks]
    answer_type_counts = answer_type_counts or {}

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = '集計レポート'
    ws.sheet_view.showGridLines = False

    THIN      = Side(style='thin', color='C0C0C0')
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    HDR_FILL  = PatternFill('solid', start_color='2E5C8A', end_color='2E5C8A')
    HDR_FONT  = Font(name='Meiryo UI', bold=True, color='FFFFFF', size=10)
    SUB_FONT  = Font(name='Meiryo UI', bold=True, size=10)
    DATA_FONT = Font(name='Meiryo UI', size=10)
    HIGH_FILL = PatternFill('solid', start_color='E2EFDA', end_color='E2EFDA')
    LOW_FILL  = PatternFill('solid', start_color='FCE4D6', end_color='FCE4D6')
    FLAG_FILL = PatternFill('solid', start_color='BDD7EE', end_color='BDD7EE')

    # カテゴリ別カラー：画面の縦棒グラフ（5.4節）と同じ色を、カテゴリ出現率の多い順に割り当てる。
    # 「中間カテゴリID」行・コード列見出し行はフル彩度、他の行はその淡色版にする。
    cat_order, cat_color_full = _category_color_map(gt, codes)
    cat_color_pale = {cid: _lighten_hex(col) for cid, col in cat_color_full.items()}
    cat_name_map   = {c['cat_id']: c['cat_name'] for c in codes}

    # 「コード別GT集計」「回答別コーディング結果」の列順（順A：カテゴリ出現率順→コード出現率順）
    cat_rank      = {cid: i for i, cid in enumerate(cat_order)}
    gt_count_map  = {g['code_id']: g['count'] for g in gt}
    codes_sorted_a = sorted(
        codes,
        key=lambda c: (cat_rank.get(c['cat_id'], len(cat_order)), -gt_count_map.get(c['code_id'], 0))
    )

    def hdr(r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font=HDR_FONT; cell.fill=HDR_FILL; cell.border=BORDER
        return cell

    def lbl(r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font=HDR_FONT; cell.fill=HDR_FILL; cell.border=BORDER
        return cell

    def dat(r, c, v, fill=None):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font=DATA_FONT; cell.border=BORDER
        if fill: cell.fill = fill
        return cell

    # ── タイトル ──────────────────────────────────────────────────
    ws.cell(row=1, column=1, value=f'集計レポート：{q_name}').font = Font(
        name='Meiryo UI', bold=True, size=14, color='2E5C8A')
    ws.cell(row=2, column=1,
            value=f'集計日時: {datetime.now().strftime("%Y/%m/%d %H:%M")}  有効回答数: {total}件').font = Font(
        name='Meiryo UI', size=10, color='808080')

    # 回答別コーディング結果の固定列：フラグ列（非該当＋不明＋無回答＋有効なリスクチェック項目）
    # ＋センチメント・回答ID・回答テキスト。「特記情報集計」「コード別GT集計」のコード列開始位置
    # （CODE_START）とも揃えるため、この並びを他のブロックでも共通して使う。
    FLAG_COLUMNS = (
        [{'key': None, 'char': '非'}]
        + [{'key': o['key'], 'char': o['char']} for o in ANSWER_TYPE_OPTIONS]
        + [{'key': o['key'], 'char': o['char']} for o in risk_opts]
    )
    FIXED_N    = 3 + len(FLAG_COLUMNS)
    CODE_START = FIXED_N + 1

    # 特記情報集計に載せる項目（センチメント＋非該当＋回答分類＋リスクチェック）。この順序が
    # 「特記情報集計」「回答別コーディング結果」のフラグ列（非/不/無/リスク項目）の並びと一致する。
    info_items = (
        [(label, sent_counts[key]) for key, label in
         [('positive', 'ポジティブ'), ('negative', 'ネガティブ'), ('neutral', 'ニュートラル'), ('mixed', '混在')]]
        + [('非該当（コードなし）', unassigned)]
        + [(o['label'], answer_type_counts.get(o['key'], 0)) for o in ANSWER_TYPE_OPTIONS]
        + [(o['label'], risk_counts.get(o['key'], 0)) for o in risk_opts]
    )

    # ── 特記情報集計（センチメント＋非該当＋回答分類＋リスクチェック） ──────
    # コード列（CODE_START）の開始位置に合わせ、GT集計・回答別コーディング結果と縦の位置を揃える。
    ws.cell(row=4, column=1, value='■ 特記情報集計').font = SUB_FONT
    for i, (label, cnt) in enumerate(info_items):
        col = CODE_START + i*2
        pct = cnt / total * 100 if total > 0 else 0
        hdr(5, col, label)
        dat(5, col+1, cnt)
        hdr(6, col, '%')
        dat(6, col+1, round(pct, 1))

    # ── GT集計（転置レイアウト） ──────────────────────────────────
    GT_START  = 8
    ws.cell(row=GT_START, column=1, value='■ コード別GT集計').font = SUB_FONT

    gt_labels = [
        'カテゴリー出現率順位',
        'カテゴリ出現数',
        '中間カテゴリID',
        '中間カテゴリ名',
        'コードID',
        'コード名',
        '出現件数',
        '出現率(%)',
        '定義',
        'ポジティブ件数',
        'ネガティブ件数',
        'ニュートラル件数',
    ]
    CAT_ID_ROW_INDEX = gt_labels.index('中間カテゴリID')  # フル彩度で塗る行
    PCT_ROW_INDEX    = gt_labels.index('出現率(%)')
    # 行見出し列はFIXED_N（回答別コーディング結果の「回答テキスト」列と同じ位置）に置き、
    # コード列（CODE_START）の開始位置を他ブロックと縦に揃える。
    for i, label in enumerate(gt_labels):
        r = GT_START + 1 + i
        lbl(r, FIXED_N, label)
        ws.row_dimensions[r].height = 18

    # カテゴリ出現数を集計
    result_map  = {r['id']: r for r in results}
    cat_counts  = {}
    code_sent   = {c['code_id']: {'positive':0,'negative':0,'neutral':0} for c in codes}
    for res in results:
        sent = res.get('sentiment', 'neutral')
        assigned_cats = set()
        for cid in res.get('codes', []):
            code = next((c for c in codes if c['code_id']==cid), None)
            if code:
                cat = code['cat_id']
                if cat not in assigned_cats:
                    cat_counts[cat] = cat_counts.get(cat, 0) + 1
                    assigned_cats.add(cat)
                if sent in code_sent[cid]:
                    code_sent[cid][sent] += 1

    # コードデータ書き込み（列順は順A：カテゴリ出現率順→コード出現率順）
    for ci, code in enumerate(codes_sorted_a):
        col  = CODE_START + ci
        cnt  = next((r['count'] for r in gt if r['code_id']==code['code_id']), 0)
        pct  = cnt / total * 100 if total > 0 else 0
        cs   = code_sent.get(code['code_id'], {'positive':0,'negative':0,'neutral':0})
        full = cat_color_full.get(code['cat_id'], 'FFFFFF')
        pale = cat_color_pale.get(code['cat_id'], 'FFFFFF')
        vals = [
            cat_rank.get(code['cat_id'], len(cat_order)) + 1,
            cat_counts.get(code['cat_id'], 0),
            code['cat_id'],
            code['cat_name'],
            code['code_id'],
            code['code_name'],
            cnt,
            round(pct, 1),
            code.get('definition', ''),
            cs['positive'],
            cs['negative'],
            cs['neutral'],
        ]
        for ri, val in enumerate(vals):
            r = GT_START + 1 + ri
            c = ws.cell(row=r, column=col, value=val)
            c.font=DATA_FONT; c.border=BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top')
            row_color = full if ri == CAT_ID_ROW_INDEX else pale
            c.fill = PatternFill('solid', start_color=row_color, end_color=row_color)
            if ri == PCT_ROW_INDEX:  # 出現率
                if pct >= 20:
                    c.fill = HIGH_FILL
                elif pct <= 2:
                    c.fill = LOW_FILL
                    c.font = Font(name='Meiryo UI', size=10, color='FF0000', bold=True)
        ws.column_dimensions[get_column_letter(col)].width = 12

    # ── 回答別コーディング結果 ────────────────────────────────────
    # 列順：フラグ列（非/不/無/リスク項目）→ センチメント → 回答ID → 回答テキスト → コード列。
    RESULT_START = GT_START + len(gt_labels) + 2
    ws.cell(row=RESULT_START, column=1, value='■ 回答別コーディング結果').font = SUB_FONT

    hdr_row = RESULT_START + 1
    META_START = len(FLAG_COLUMNS) + 1  # センチメント列の開始位置

    # フラグ列見出し（非該当・不明・無回答・有効化されたリスクチェック項目、一文字見出し）
    for ci, fc in enumerate(FLAG_COLUMNS):
        col = 1 + ci
        hdr(hdr_row, col, fc['char'])
        ws.column_dimensions[get_column_letter(col)].width = 4

    for ci, h in enumerate(['センチメント', '回答ID', '回答テキスト']):
        hdr(hdr_row, META_START+ci, h)

    for ci, code in enumerate(codes_sorted_a):
        col  = CODE_START + ci
        full = cat_color_full.get(code['cat_id'], 'FFFFFF')
        c = ws.cell(row=hdr_row, column=col, value=code['code_id'])
        c.font = Font(name='Meiryo UI', bold=True, size=10, color='000000')
        c.fill = PatternFill('solid', start_color=full, end_color=full); c.border=BORDER

    item_map = {item['id']: item['text'] for item in items}
    for ri, res in enumerate(results):
        r        = hdr_row + 1 + ri
        rid      = res.get('id', '')
        text     = item_map.get(rid, '')[:50]
        sent     = res.get('sentiment', '')
        assigned = res.get('codes', [])
        for ci, val in enumerate([sent, rid, text]):
            dat(r, META_START+ci, val)
        for ci, fc in enumerate(FLAG_COLUMNS):
            flagged = (not assigned) if fc['key'] is None else bool(res.get(fc['key']))
            c = ws.cell(row=r, column=1+ci, value=1 if flagged else '')
            c.font=DATA_FONT; c.border=BORDER
            if flagged: c.fill = FLAG_FILL
        for ci, code in enumerate(codes_sorted_a):
            col  = CODE_START + ci
            flag = 1 if code['code_id'] in assigned else 0
            pale = cat_color_pale.get(code['cat_id'], 'FFFFFF')
            c = ws.cell(row=r, column=col, value=flag if flag else '')
            c.font=DATA_FONT; c.border=BORDER
            c.fill = PatternFill('solid', start_color=pale, end_color=pale)
            if flag: c.fill = FLAG_FILL

    ws.column_dimensions[get_column_letter(META_START)].width   = 12  # センチメント
    ws.column_dimensions[get_column_letter(META_START+1)].width = 10  # 回答ID
    ws.column_dimensions[get_column_letter(META_START+2)].width = 30  # 回答テキスト
    ws.freeze_panes = f'A{hdr_row+1}'

    # オートフィルター（見出し行〜最終回答行、全列に設定）
    last_col = CODE_START + len(codes) - 1 if codes else FIXED_N
    ws.auto_filter.ref = f'A{hdr_row}:{get_column_letter(last_col)}{hdr_row + len(results)}'

    # ══════════════════════════════════════════════════
    # 集計ダイジェストシート
    # ══════════════════════════════════════════════════
    ws2 = wb.create_sheet('集計ダイジェスト')
    ws2.sheet_view.showGridLines = False

    # タイトル
    ws2.cell(row=1, column=1, value=f'After Coder 集計ダイジェスト：{q_name}').font = Font(
        name='Meiryo UI', bold=True, size=14, color='2E5C8A')
    ws2.cell(row=2, column=1,
             value=f'集計日時: {datetime.now().strftime("%Y/%m/%d %H:%M")}  有効回答数: {total}件').font = Font(
        name='Meiryo UI', size=10, color='808080')

    # 特記情報集計（ws側で構築したinfo_itemsをそのまま再利用し、項目・並び順を一致させる）
    ws2.cell(row=4, column=1, value='■ 特記情報集計').font = SUB_FONT
    for i, (label, cnt) in enumerate(info_items):
        col = 1 + i*2
        pct = cnt / total * 100 if total > 0 else 0
        c = ws2.cell(row=5, column=col, value=label)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.border=BORDER
        c2 = ws2.cell(row=5, column=col+1, value=cnt)
        c2.font=DATA_FONT; c2.border=BORDER
        c3 = ws2.cell(row=6, column=col, value='%')
        c3.font=HDR_FONT; c3.fill=HDR_FILL; c3.border=BORDER
        c4 = ws2.cell(row=6, column=col+1, value=round(pct,1))
        c4.font=DATA_FONT; c4.border=BORDER

    # GT集計（行方向・出現率順）
    ws2.cell(row=8, column=1, value='■ コード別GT集計').font = SUB_FONT
    gt_headers = ['カテゴリID', 'カテゴリ名', 'コードID', 'コード名', '出現件数', '出現率(%)', '定義', 'キーワード']
    for ci, h in enumerate(gt_headers):
        c = ws2.cell(row=9, column=ci+1, value=h)
        c.font=HDR_FONT; c.fill=HDR_FILL; c.border=BORDER

    code_kw_map = {}
    for cd in codes:
        kw = cd.get('keywords', [])
        if isinstance(kw, list):
            kw = '; '.join(kw)
        code_kw_map[cd['code_id']] = kw or ''

    # 出現率順にソート
    gt_sorted = sorted(gt, key=lambda x: x['count'], reverse=True)
    for ri, row_data in enumerate(gt_sorted):
        r    = 10 + ri
        pct  = row_data['pct']
        pale = cat_color_pale.get(row_data['cat_id'], 'FFFFFF')
        vals = [
            row_data['cat_id'],
            row_data['cat_name'],
            row_data['code_id'],
            row_data['code_name'],
            row_data['count'],
            pct,
            row_data['definition'],
            code_kw_map.get(row_data['code_id'], ''),
        ]
        for ci, val in enumerate(vals):
            c = ws2.cell(row=r, column=ci+1, value=val)
            c.font=DATA_FONT; c.border=BORDER
            c.fill = PatternFill('solid', start_color=pale, end_color=pale)
            if ci == 5:  # 出現率列
                if pct >= 20:
                    c.fill = HIGH_FILL
                elif pct <= 2:
                    c.fill = LOW_FILL
                    c.font = Font(name='Meiryo UI', size=10, color='FF0000', bold=True)

    # 列幅
    for ci, w in enumerate([10, 22, 10, 24, 10, 10, 40, 30]):
        ws2.column_dimensions[get_column_letter(ci+1)].width = w
    ws2.freeze_panes = 'A10'

    # ── 棒グラフ用の隠しヘルパー表 ──────────────────────────────
    # Excelのネイティブな棒グラフには「1系列内で棒ごとに個別の色・凡例エントリを持たせる」機能が
    # ないため（openpyxlのDataPoint個別着色は色は付くが凡例が棒の数だけ増えてしまい、
    # カテゴリ単位の凡例にならない）、カテゴリごとに列を分けた表を裏で作り、
    # カテゴリ数ぶんの系列（＝カテゴリごとに固定色・凡例1個）として描画する。
    # 該当しない行は空欄にすることで、各コードの位置に実際には1本しか棒が出ないようにする
    # （chart.overlap=100と合わせて完全に重ねる）。可視の集計表（上記）とは順序が異なる
    # （こちらは「操作画面の順A＝カテゴリ出現率順→コード出現率順」に固定）。
    cat_rank    = {cid: i for i, cid in enumerate(cat_order)}
    gt_sorted_a = sorted(gt, key=lambda x: (cat_rank.get(x['cat_id'], len(cat_order)), -x['count']))

    if gt_sorted_a:
        HELPER_START_COL = 10  # J列から（表示テーブルはA〜H列なので余裕を持たせる）
        helper_label_col = HELPER_START_COL
        chart_hdr_row    = 9
        chart_row_start  = 10
        chart_row_end    = chart_row_start + len(gt_sorted_a) - 1

        ws2.cell(row=chart_hdr_row, column=helper_label_col, value='コード名（グラフ用・カテゴリ順）')
        for i, cid in enumerate(cat_order):
            ws2.cell(row=chart_hdr_row, column=helper_label_col + 1 + i, value=cat_name_map.get(cid, cid))
        for ri, row_data in enumerate(gt_sorted_a):
            r = chart_row_start + ri
            ws2.cell(row=r, column=helper_label_col, value=row_data['code_name'])
            for i, cid in enumerate(cat_order):
                col = helper_label_col + 1 + i
                ws2.cell(row=r, column=col, value=row_data['pct'] if row_data['cat_id'] == cid else None)

        helper_last_col = helper_label_col + len(cat_order)
        for c in range(helper_label_col, helper_last_col + 1):
            ws2.column_dimensions[get_column_letter(c)].hidden = True

        # 棒グラフ（画面の「コード別GT集計」と同じ配色・同じ並び順＝カテゴリ出現率順→コード出現率順）
        chart = BarChart()
        chart.type     = 'col'
        chart.grouping = 'clustered'
        chart.overlap  = 100  # 系列を完全に重ね、コードごとに実質1本の棒に見せる
        chart.gapWidth = 50
        chart.title = 'コード別出現率(%)（カテゴリ出現率順→コード出現率順）'
        chart.y_axis.title = '出現率(%)'
        # データ元（J列以降）を非表示列にしているため、Excelの既定「表示セルのみプロット」設定
        # (plotVisOnly=True)のままだと非表示列のデータが一切プロットされず、グラフが空になる。
        chart.visible_cells_only = False
        cats = Reference(ws2, min_col=helper_label_col, min_row=chart_row_start, max_row=chart_row_end)
        for i, cid in enumerate(cat_order):
            col  = helper_label_col + 1 + i
            data = Reference(ws2, min_col=col, min_row=chart_hdr_row, max_row=chart_row_end)
            chart.add_data(data, titles_from_data=True)
        # chart.set_categories()は参照先の型に関わらず常にNumRefとして埋め込むため、
        # コード名（文字列）を渡すとExcel側でX軸の項目名が表示されない。文字列参照として
        # 明示するため、StrRefを直接組み立ててseries.catに設定する。
        cats_str = f'{cats}'
        for series in chart.series:
            series.cat = AxDataSource(strRef=StrRef(f=cats_str))
        for i, series in enumerate(chart.series):
            cid = cat_order[i]
            series.graphicalProperties = GraphicalProperties(solidFill=cat_color_full.get(cid, 'FFFFFF'))
        chart.width  = 26
        chart.height = 11
        chart_anchor_row = chart_row_end + 3
        ws2.add_chart(chart, f'A{chart_anchor_row}')

        # ── グラフ下の集計表（グラフと同じ順＝順A：カテゴリ出現率順→コード出現率順） ──
        # チャートの高さ(11cm)は既定の行高（約0.53cm/行）でおよそ21行分にあたるため、
        # 重ならないよう十分な余白を空けて配置する。
        TABLE_START = chart_anchor_row + 23
        ws2.cell(row=TABLE_START, column=1, value='■ コード別集計表（グラフと同じ順）').font = SUB_FONT
        table_hdr_row = TABLE_START + 1
        row_labels = ['コード名', 'コード出現率(%)', 'コード出現数']
        for i, label in enumerate(row_labels):
            c = ws2.cell(row=table_hdr_row + i, column=1, value=label)
            c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
        for ci, row_data in enumerate(gt_sorted_a):
            col  = 2 + ci
            full = cat_color_full.get(row_data['cat_id'], 'FFFFFF')
            for ri, val in enumerate([row_data['code_name'], row_data['pct'], row_data['count']]):
                c = ws2.cell(row=table_hdr_row + ri, column=col, value=val)
                c.font = DATA_FONT; c.border = BORDER
                c.fill = PatternFill('solid', start_color=full, end_color=full)
            ws2.column_dimensions[get_column_letter(col)].width = 12

    # ══════════════════════════════════════════════════
    # 自由回答一覧シート（カテゴリー→コード別に実例を並べた一覧）
    # ══════════════════════════════════════════════════
    ws3 = wb.create_sheet('自由回答一覧')
    ws3.sheet_view.showGridLines = False

    ws3.cell(row=1, column=1, value=f'自由回答一覧：{q_name}').font = Font(
        name='Meiryo UI', bold=True, size=16, color='2E5C8A')
    ws3.cell(row=2, column=1,
             value=f'集計日時: {datetime.now().strftime("%Y/%m/%d %H:%M")}  有効回答数: {total}件').font = Font(
        name='Meiryo UI', size=10, color='808080')

    item_map_full = {it['id']: it for it in items}
    code_lookup   = {c['code_id']: c for c in codes}
    code_rank     = {c['code_id']: i for i, c in enumerate(codes_sorted_a)}

    # 属性列名：実際に使われた属性キーを、最初に登場した順に確定する（列数・列名は可変）
    attr_keys = []
    for it in items:
        for k in it.get('attrs', {}):
            if k not in attr_keys:
                attr_keys.append(k)

    def _attr_sort_key(it):
        return tuple(str(it.get('attrs', {}).get(k, '')) for k in attr_keys)

    def _excerpt_for_code(res, cid, fallback_text):
        """
        resの文単位データ（sentences、5.1.1節参照）から、該当コードcidが付与された文だけを
        抜き出して連結する（原文の一部をそのまま引用するため、要約や言い換えは発生しない）。
        文単位データが無い場合（旧形式の結果など）は回答全文にフォールバックする。
        これにより、1回答が複数コードに該当する場合でも、コードごとに関係する箇所だけを
        表示でき、無関係なコードの実例として全文がそのまま出てくることを防ぐ。

        該当文が複数ある場合は全て連結して表示する（1文に絞る案も試したが、同じコードが
        複数文にまたがるケースでは、経緯や状況の流れが1文だけでは伝わらずかえって分かり
        にくいというユーザーの判断により、全文連結の方針を維持している）。
        """
        sentences = res.get('sentences') or []
        matched = [s.get('text', '') for s in sentences if cid in s.get('codes', []) and s.get('text')]
        return '／'.join(matched) if matched else fallback_text

    # (順A用のランク, 属性値の並び, 元の出現順, コード情報, 回答item, 該当箇所の抜粋) のリストを作り、
    # 非該当は別枠に分ける
    coded_rows, unassigned_items = [], []
    for ri, res in enumerate(results):
        it = item_map_full.get(res.get('id'))
        if not it:
            continue
        assigned = res.get('codes', [])
        if not assigned:
            unassigned_items.append(it)
            continue
        for cid in assigned:
            code = code_lookup.get(cid)
            if code:
                excerpt = _excerpt_for_code(res, cid, it['text'])
                coded_rows.append((code_rank.get(cid, len(codes_sorted_a)), _attr_sort_key(it), ri, code, it, excerpt))
    # 順A（カテゴリ→コード）を主キーに、属性１・属性２・属性３…の値を副キーとして並べる
    coded_rows.sort(key=lambda x: (x[0], x[1], x[2]))

    # 列構成：A列＝カテゴリー/コードの見出し専用（データは持たない）、B列以降が実データ
    MARK_COL = 1
    DATA_START = 2
    headers = ['カテゴリー名', 'コード名', '回答者ID', 'FA番号', '自由記述（該当箇所）'] + attr_keys
    hdr_row3 = 4
    hdr_mark = ws3.cell(row=hdr_row3, column=MARK_COL)
    hdr_mark.font = HDR_FONT; hdr_mark.fill = HDR_FILL; hdr_mark.border = BORDER
    hdr_mark.alignment = Alignment(wrap_text=False, vertical='center')
    for ci, h in enumerate(headers):
        c = ws3.cell(row=hdr_row3, column=DATA_START+ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(wrap_text=False, vertical='center')

    last_col = DATA_START + len(headers) - 1

    def _mark_cell(r, text, color, size, font_color='FFFFFF'):
        """カテゴリー／コードの切り替わりを示す見出しテキストはA列のみに入れるが、
        背景色は行の区切りとして視認しやすいよう最終列まで塗る（セルは結合しない）"""
        for col in range(MARK_COL, last_col + 1):
            c = ws3.cell(row=r, column=col, value=text if col == MARK_COL else None)
            c.font = Font(name='Meiryo UI', bold=True, size=size, color=font_color)
            c.fill = PatternFill('solid', start_color=color, end_color=color)
            c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top')

    def _data_row(r, vals):
        for ci, v in enumerate(vals):
            c = ws3.cell(row=r, column=DATA_START+ci, value=v)
            c.font = DATA_FONT; c.border = BORDER
            c.alignment = Alignment(wrap_text=True, vertical='top')

    r = hdr_row3 + 1
    prev_cat = prev_code = None
    for _, _, _, code, it, excerpt in coded_rows:
        cat_id = code['cat_id']
        if cat_id != prev_cat:
            _mark_cell(r, f"■ カテゴリー：{code['cat_name']}", cat_color_full.get(cat_id, '808080'), size=14)
            r += 1
            prev_cat, prev_code = cat_id, None
        if code['code_id'] != prev_code:
            _mark_cell(r, f"▶ コード：{code['code_name']}", cat_color_pale.get(cat_id, 'D9D9D9'), size=12, font_color='000000')
            r += 1
            prev_code = code['code_id']
        _data_row(r, [code['cat_name'], code['code_name'], it['id'], it.get('fa_no') or '', excerpt]
                  + [it.get('attrs', {}).get(k, '') for k in attr_keys])
        r += 1

    if unassigned_items:
        _mark_cell(r, '■ 非該当（コードなし）', '808080', size=14)
        r += 1
        for it in sorted(unassigned_items, key=_attr_sort_key):
            _data_row(r, ['', '', it['id'], it.get('fa_no') or '', it['text']]
                      + [it.get('attrs', {}).get(k, '') for k in attr_keys])
            r += 1

    ws3.column_dimensions[get_column_letter(MARK_COL)].width = 26
    ws3.column_dimensions[get_column_letter(DATA_START)].width   = 20  # カテゴリー名
    ws3.column_dimensions[get_column_letter(DATA_START+1)].width = 25  # コード名
    ws3.column_dimensions[get_column_letter(DATA_START+2)].width = 12  # 回答者ID
    ws3.column_dimensions[get_column_letter(DATA_START+3)].width = 10  # FA番号
    ws3.column_dimensions[get_column_letter(DATA_START+4)].width = 50  # 自由記述
    for ci in range(len(attr_keys)):
        ws3.column_dimensions[get_column_letter(DATA_START+5+ci)].width = 14
    ws3.freeze_panes = f'{get_column_letter(DATA_START)}{hdr_row3+1}'

    # ══════════════════════════════════════════════════
    # 文単位データシート（1行1文のRAWデータ。他の集計・分析への再利用を想定）
    # ══════════════════════════════════════════════════
    ws4 = wb.create_sheet('文単位データ')
    ws4.sheet_view.showGridLines = False

    ws4.cell(row=1, column=1, value=f'文単位データ：{q_name}').font = Font(
        name='Meiryo UI', bold=True, size=16, color='2E5C8A')
    ws4.cell(row=2, column=1,
             value=f'集計日時: {datetime.now().strftime("%Y/%m/%d %H:%M")}  有効回答数: {total}件').font = Font(
        name='Meiryo UI', size=10, color='808080')
    ws4.cell(row=3, column=1,
             value='※回答を文単位に分割し、文ごとのセンチメント・該当コードをそのまま列挙したRAWデータです。'
                   '他の集計・分析に再利用する用途を想定しています。').font = Font(
        name='Meiryo UI', size=9, italic=True, color='808080')

    sent4_headers = ['回答ID', 'FA番号'] + attr_keys + ['文番号', '文テキスト', 'センチメント', '該当コードID', '該当コード名']
    hdr_row4 = 5
    for ci, h in enumerate(sent4_headers):
        c = ws4.cell(row=hdr_row4, column=1+ci, value=h)
        c.font = HDR_FONT; c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(wrap_text=False, vertical='center')

    r4 = hdr_row4 + 1
    for res in results:
        it = item_map_full.get(res.get('id'))
        if not it:
            continue
        sentences = res.get('sentences') or []
        for s in sentences:
            code_ids   = s.get('codes', [])
            code_names = [code_lookup[cid]['code_name'] for cid in code_ids if cid in code_lookup]
            vals = (
                [it['id'], it.get('fa_no') or '']
                + [it.get('attrs', {}).get(k, '') for k in attr_keys]
                + [s.get('idx', ''), s.get('text', ''), s.get('sentiment', ''),
                   '; '.join(code_ids), '; '.join(code_names)]
            )
            for ci, v in enumerate(vals):
                c = ws4.cell(row=r4, column=1+ci, value=v)
                c.font = DATA_FONT; c.border = BORDER
                c.alignment = Alignment(wrap_text=True, vertical='top')
            r4 += 1

    id_col_count = 2 + len(attr_keys)
    ws4.column_dimensions[get_column_letter(1)].width = 12  # 回答ID
    ws4.column_dimensions[get_column_letter(2)].width = 10  # FA番号
    for ci in range(len(attr_keys)):
        ws4.column_dimensions[get_column_letter(3+ci)].width = 14
    ws4.column_dimensions[get_column_letter(id_col_count+1)].width = 8   # 文番号
    ws4.column_dimensions[get_column_letter(id_col_count+2)].width = 50  # 文テキスト
    ws4.column_dimensions[get_column_letter(id_col_count+3)].width = 12  # センチメント
    ws4.column_dimensions[get_column_letter(id_col_count+4)].width = 20  # 該当コードID
    ws4.column_dimensions[get_column_letter(id_col_count+5)].width = 30  # 該当コード名
    ws4.freeze_panes = f'A{hdr_row4+1}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ══════════════════════════════════════════════════
# コードブック策定方式（A/B/C）
# ══════════════════════════════════════════════════

CODEBOOK_MODES = [
    {'code': 'A',        'label': '方式A：標準'},
    {'code': 'B',        'label': '方式B：トップダウン＋ボトムアップ'},
    {'code': 'C1',       'label': '方式C1：全件精査（1/4抽出）'},
    {'code': 'C2',       'label': '方式C2：全件精査（ハーフ抽出）'},
    {'code': 'C3',       'label': '方式C3：全件精査（フル抽出）'},
    {'code': 'C_AUTO',   'label': '方式C-自動：全件精査（件数自動調整）（推奨）'},
    {'code': 'EXISTING', 'label': '既存のコードブックを使用（アップロード）'},
]
CODEBOOK_MODE_LABELS  = [m['label'] for m in CODEBOOK_MODES]
CODEBOOK_MODE_CODE    = {m['label']: m['code'] for m in CODEBOOK_MODES}
CODEBOOK_MODE_C_RATIO = {'C1': 0.25, 'C2': 0.5, 'C3': 1.0}


def _auto_sample_ratio(n):
    """
    方式C-自動のStage1抽出比率。抽出数f(N) = min(N, 10√N, 1000)をNで割った比率を返す。
    - N≤100：全部読む（f(N)=Nとなるのは N≤100 のときのみ、10√N≥Nと同値）
    - N>100：10√Nで緩やかに増やす（サンプル比率はNが大きいほど下がる）
    - N>10,000：10√N≥1000となるため1000件で頭打ち
    """
    if n <= 0:
        return 1.0
    target = min(n, 10 * math.sqrt(n), 1000)
    return target / n


def _diff_detect_loop(client, codebook, remaining, max_codes, q_name, progress_bar, p_start=0.05, p_end=0.30):
    """既存コードブックに対し残りサンプルをバッチ処理し新規コードを差分検出する（方式A・Bで共用）"""
    total_batches = max(len(remaining) // 20, 1)
    round_no = 2
    while remaining:
        batch     = remaining[:20]
        remaining = remaining[20:]
        new_list  = llm_detect_new_codes(client, batch, codebook, max_codes, q_name)
        if new_list:
            for nc in new_list:
                for cat in codebook['categories']:
                    if cat['cat_id'] == nc.get('cat_id'):
                        cat['codes'].append({
                            'code_id':    nc.get('code_id', ''),
                            'code_name':  nc.get('code_name', ''),
                            'definition': nc.get('definition', ''),
                        })
                        break
        cur = sum(len(c['codes']) for c in codebook.get('categories', []))
        if cur >= max_codes:
            break
        progress_bar.progress(min(p_start + (p_end - p_start) * (round_no / total_batches), p_end))
        round_no += 1
    return codebook


def _build_codebook_a(client, all_items, max_codes, q_name, data_context, progress_bar):
    """方式A：標準 - 少量サンプルでコードブックを生成し差分検出"""
    sample_1 = all_items[:30]
    codebook = llm_generate_codebook(client, sample_1, max_codes, q_name, data_context)
    if not codebook:
        return None
    return _diff_detect_loop(client, codebook, all_items[30:], max_codes, q_name, progress_bar)


def _build_codebook_b(client, all_items, max_codes, q_name, data_context, progress_bar, status_text):
    """方式B：トップダウン＋ボトムアップ - 骨格を先に生成し実データで具体化・差分検出"""
    if not data_context.strip():
        status_text.markdown('**Step 1/3** 「分析データの特徴」が未入力のため方式Aで生成します...')
        return _build_codebook_a(client, all_items, max_codes, q_name, data_context, progress_bar)

    status_text.markdown('**Step 1/3** トップダウンで骨格を生成中...')
    skeleton = llm_generate_codebook_topdown(client, data_context, q_name, max_codes)
    if not skeleton:
        return None
    progress_bar.progress(0.10)

    status_text.markdown('**Step 1/3** 実データで骨格を具体化中...')
    codebook = llm_elaborate_skeleton(client, skeleton, all_items[:30], max_codes, q_name, data_context)
    if not codebook:
        return None
    progress_bar.progress(0.20)

    return _diff_detect_loop(client, codebook, all_items[30:], max_codes, q_name, progress_bar, p_start=0.20, p_end=0.30)


def _topics_cache_key(all_items, q_name, data_context):
    """方式Cのキャッシュキー。回答内容・プロジェクト名・分析データの特徴が同じなら同一キーになる（回答順序は無視）"""
    texts = sorted(x['text'] for x in all_items)
    src = q_name + '␟' + data_context + '␟' + '␞'.join(texts)
    return hashlib.sha256(src.encode('utf-8')).hexdigest()


# Stage1主題抽出の進捗をディスクに保存するディレクトリ。
# st.session_stateはアプリ再起動・セッション切断で消えてしまうため、
# 大量データ（全件精査など）でバッチ数が多い場合に途中経過を失わないよう、
# ディスクにも進捗を残し、次回同じデータで実行した際に未処理バッチから再開できるようにする。
STAGE1_CACHE_DIR = APP_DIR / '.codebook_cache'


def _stage1_cache_path(cache_key):
    # cache_keyには「:」（ratioの区切り）が含まれ、Windowsのファイル名では
    # 特殊な意味を持つ（NTFSの代替データストリームとして扱われ、ファイルが
    # 実質空になる）ため、ファイル名にはcache_key自体ではなく再ハッシュした値を使う。
    safe_name = hashlib.sha256(cache_key.encode('utf-8')).hexdigest()
    return STAGE1_CACHE_DIR / f'{safe_name}.json'


def _load_stage1_checkpoint(cache_key):
    path = _stage1_cache_path(cache_key)
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except (OSError, json.JSONDecodeError):
        return None


def _save_stage1_checkpoint(cache_key, data):
    STAGE1_CACHE_DIR.mkdir(exist_ok=True)
    _stage1_cache_path(cache_key).write_text(json.dumps(data, ensure_ascii=False), encoding='utf-8')


def _clear_stage1_checkpoint(cache_key):
    path = _stage1_cache_path(cache_key)
    if path.exists():
        path.unlink()


def _cleanup_old_stage1_checkpoints(max_age_hours=24):
    """古い（放棄された）チェックポイントファイルを掃除する。実行のたびに軽く呼ぶだけの簡易処理。"""
    if not STAGE1_CACHE_DIR.exists():
        return
    cutoff = time.time() - max_age_hours * 3600
    for f in STAGE1_CACHE_DIR.glob('*.json'):
        try:
            if f.stat().st_mtime < cutoff:
                f.unlink()
        except OSError:
            pass


def _build_codebook_c_stage1(client, stage1_items, max_codes, q_name, data_context, progress_bar, status_text,
                              cache_key):
    """
    方式C Stage1：ランダム抽出した一部から主題抽出→統合し初期コードブックを作る。
    大量データ（全件精査など）では主題抽出のバッチ数が多く1回の実行時間が長くなるため、
    バッチ完了ごとに進捗をディスクへ保存する。アプリの再起動やセッション切断で処理が
    中断しても、同じデータで再実行すれば未処理バッチから再開でき、最初からやり直しにならない。
    """
    _cleanup_old_stage1_checkpoints()
    batches = [stage1_items[i:i+50] for i in range(0, len(stage1_items), 50)]

    checkpoint = _load_stage1_checkpoint(cache_key)
    if checkpoint and checkpoint.get('total_batches') == len(batches):
        topics_all         = checkpoint.get('topics_all', [])
        completed_batches  = checkpoint.get('completed_batches', 0)
    else:
        topics_all        = []
        completed_batches = 0

    if completed_batches > 0:
        status_text.markdown(
            f'**Step 1/3** Stage1: 前回の続きから再開します（{completed_batches}/{len(batches)}バッチ済み）...'
        )

    for bi in range(completed_batches, len(batches)):
        status_text.markdown(f'**Step 1/3** Stage1（{len(stage1_items)}件）: 主題抽出中... {bi+1}/{len(batches)}バッチ')
        topics_all.extend(llm_extract_topics(client, batches[bi], q_name, data_context))
        progress_bar.progress(min(0.05 + 0.08 * ((bi+1) / len(batches)), 0.13))
        _save_stage1_checkpoint(cache_key, {
            'total_batches':     len(batches),
            'completed_batches': bi + 1,
            'topics_all':        topics_all,
        })

    topics_dedup = list(dict.fromkeys(topics_all))

    # 主題数が多いと統合コールが一括で処理しきれず失敗しやすいため、
    # チャンクごとにLLMで類似表現をまとめ、段階的に件数を減らしてから統合に渡す
    REDUCE_THRESHOLD = 300
    REDUCE_CHUNK = 200
    reduce_round = 1
    while len(topics_dedup) > REDUCE_THRESHOLD:
        status_text.markdown(f'**Step 1/3** Stage1: 主題リストを整理中...（{len(topics_dedup)}件、{reduce_round}回目）')
        chunks  = [topics_dedup[i:i+REDUCE_CHUNK] for i in range(0, len(topics_dedup), REDUCE_CHUNK)]
        reduced = []
        for chunk in chunks:
            reduced.extend(llm_reduce_topics(client, chunk, q_name, data_context))
        new_dedup = list(dict.fromkeys(reduced))
        if len(new_dedup) >= len(topics_dedup):
            break  # これ以上減らない場合は打ち切り、現状のリストで統合に進む
        topics_dedup = new_dedup
        reduce_round += 1

    status_text.markdown(
        f'**Step 1/3** Stage1: 初期コードブックを確定中...（{len(topics_all)}件 → 整理後{len(topics_dedup)}件）'
    )
    codebook = llm_consolidate_topics(client, topics_dedup, max_codes, q_name, data_context)
    progress_bar.progress(0.15)
    if codebook:
        # 成功した場合のみチェックポイントを消す。失敗時は残しておき、
        # 再実行時に抽出済みの主題リストから（再抽出せず）統合からやり直せるようにする。
        _clear_stage1_checkpoint(cache_key)
    return codebook


def _build_codebook_c(client, all_items, max_codes, q_name, data_context, progress_bar, status_text, ratio=0.25):
    """
    方式C1/C2/C3共通処理：全件精査
    Stage1: 全体のratio割合をランダム抽出して主題抽出→統合し初期コードブックを作成
    Stage2: 残り（1-ratio）を差分検出（ratio=1.0の場合は残りがないため実施しない）
    ステージ完了ごとの結果はセッション内でキャッシュし、後段の失敗時に前段からの
    やり直しを避ける（同一データ・プロジェクト名・分析データの特徴・ratioの場合のみ再利用）。
    さらにStage1内部（主題抽出）はバッチごとにディスクへも進捗を保存しており、
    大量データでStage1の途中にアプリが再起動・セッション切断しても再開できる（詳細は
    `_build_codebook_c_stage1`参照）。
    """
    cache_key   = _topics_cache_key(all_items, q_name, data_context) + f':{ratio}'
    stage_cache = st.session_state.setdefault('stage_codebook_cache', {})
    cached      = stage_cache.get(cache_key)
    stage_done  = cached['stage'] if cached else 0
    codebook    = cached['codebook'] if cached else None

    n1        = max(int(len(all_items) * ratio), 1)
    remaining = all_items[n1:]

    if stage_done >= 1:
        status_text.markdown('**Step 1/3** Stage1: キャッシュ済みの初期コードブックを再利用...')
        progress_bar.progress(0.15 if remaining else 0.30)
    else:
        codebook = _build_codebook_c_stage1(
            client, all_items[:n1], max_codes, q_name, data_context, progress_bar, status_text, cache_key
        )
        if not codebook:
            return None
        stage_done = 1
        stage_cache[cache_key] = {'stage': 1, 'codebook': codebook}

    if not remaining:
        progress_bar.progress(0.30)
        return codebook

    if stage_done >= 2:
        status_text.markdown('**Step 1/3** Stage2: キャッシュ済みの結果を再利用...')
        progress_bar.progress(0.30)
    else:
        status_text.markdown(f'**Step 1/3** Stage2: 残り{len(remaining)}件を差分検出中...')
        codebook = _diff_detect_loop(
            client, codebook, remaining, max_codes, q_name, progress_bar, p_start=0.15, p_end=0.30
        )
        stage_cache[cache_key] = {'stage': 2, 'codebook': codebook}

    return codebook


# ══════════════════════════════════════════════════
# メイン処理
# ══════════════════════════════════════════════════

def _generate_codebook_step(client, all_items, max_codes, q_name, data_context, progress_bar, status_text,
                             codebook_mode, existing_codebook=None):
    """Step1相当：指定方式でコードブックを生成（または既存コードブックを使用）して返す"""
    status_text.markdown('**コードブックを生成中...**')
    progress_bar.progress(0.05)

    if codebook_mode == 'EXISTING':
        status_text.markdown('**アップロードされたコードブックを使用します**')
        codebook = existing_codebook
        progress_bar.progress(0.30)
        return codebook

    if codebook_mode == 'B':
        codebook = _build_codebook_b(client, all_items, max_codes, q_name, data_context, progress_bar, status_text)
    elif codebook_mode == 'C_AUTO':
        ratio = _auto_sample_ratio(len(all_items))
        codebook = _build_codebook_c(client, all_items, max_codes, q_name, data_context, progress_bar, status_text, ratio)
    elif codebook_mode in CODEBOOK_MODE_C_RATIO:
        ratio = CODEBOOK_MODE_C_RATIO[codebook_mode]
        codebook = _build_codebook_c(client, all_items, max_codes, q_name, data_context, progress_bar, status_text, ratio)
    else:
        codebook = _build_codebook_a(client, all_items, max_codes, q_name, data_context, progress_bar)

    # プロンプトの「総数N個以内」はあくまで指示であり、LLMが厳密に守るとは限らない
    # （実際に上限を超えるケースが確認された）。生成直後に総数を確認し、超えていれば
    # 既存の編集機能（llm_edit_codebook）を使って統合し直す安全弁を設ける。
    if codebook:
        total = sum(len(cat.get('codes', [])) for cat in codebook.get('categories', []))
        if total > max_codes:
            status_text.markdown(f'**コード数が上限（{max_codes}個）を超えています（{total}個）。統合中...**')
            instruction = (
                f'コード総数が{total}個あり、上限の{max_codes}個を超えています。'
                f'内容が近いコード同士を統合し、必ず総数{max_codes}個以内に収めてください（絶対に超えないこと）。'
                f'カテゴリ構成やコードIDの命名規則は維持してください。'
            )
            trimmed = llm_edit_codebook(client, codebook, instruction, q_name)
            if trimmed and trimmed.get('categories'):
                codebook = trimmed

    return codebook


CODING_SCOPE_OPTIONS = ['100件でコーディング', '200件でコーディング', '全件コーディング']
CODING_SCOPE_SIZES   = {
    '100件でコーディング':          100,
    '200件でコーディング':          200,
    '全件コーディング':              None,  # Noneは実行時に全件数へ解決
}


CODING_BATCH_SIZE = 15


def _code_one_batch(client, q_name, codes, batch, coding_model, enabled_risks, coding_strictness):
    """1バッチ分をコーディングし、無回答（空欄）のルールベース判定を上書きして返す"""
    res = llm_code_batch(client, batch, codes, q_name, model=coding_model, enabled_risks=enabled_risks,
                         strictness=coding_strictness)
    # 無回答（空欄）はAIに判定させず、回答テキストから機械的に確定させる。
    # 空欄なのに「不明」も立ってしまうと二重計上になるため、その場合はunclearを上書きする。
    batch_text = {x['id']: x['text'] for x in batch}
    for r in res:
        unanswered = not batch_text.get(r.get('id'), '').strip()
        r['unanswered'] = unanswered
        if unanswered:
            r['unclear'] = False
    return res


def _code_items(client, q_name, codes, items, progress_bar, status_text, p_start=0.30, p_end=0.90,
                 coding_model=CODING_MODEL, enabled_risks=None, coding_strictness=CODING_STRICTNESS):
    """itemsをコーディングし、結果リストを返す（集計は行わない）"""
    if not items:
        return []
    results = []
    batches = [items[i:i+CODING_BATCH_SIZE] for i in range(0, len(items), CODING_BATCH_SIZE)]
    for bi, batch in enumerate(batches):
        res = _code_one_batch(client, q_name, codes, batch, coding_model, enabled_risks, coding_strictness)
        results.extend(res)
        pct = p_start + (p_end - p_start) * ((bi+1) / len(batches))
        progress_bar.progress(min(pct, p_end))
        status_text.markdown(f'**コーディング中...** {bi+1}/{len(batches)}バッチ')
        time.sleep(0.1)
    return results


def _build_codebook_and_codes(client, all_items, max_codes, q_name, data_context, progress_bar, status_text,
                               codebook_mode, existing_codebook):
    """コードブックを策定（または既存コードブックを使用）し、(codebook, codes)を返す。失敗時は(None, None)。"""
    codebook = _generate_codebook_step(
        client, all_items, max_codes, q_name, data_context, progress_bar, status_text,
        codebook_mode, existing_codebook
    )
    if not codebook:
        reason = get_last_error() or '原因不明（AIから有効なコードブック構造が返されませんでした）'
        st.error(
            'コードブック生成に失敗しました。再度お試しください。'
            + f'\n\n詳細: {reason}'
        )
        return None, None

    codes = [
        {**c, 'cat_id': cat['cat_id'], 'cat_name': cat['cat_name']}
        for cat in codebook.get('categories', [])
        for c in cat.get('codes', [])
    ]
    status_text.markdown(f'**コードブック完成：{len(codes)}コード ✓**')
    progress_bar.progress(0.30)
    return codebook, codes


def run_pipeline(api_key, q_name, items, max_codes, progress_bar, status_text, data_context='',
                  codebook_mode='A', existing_codebook=None, sample_size=None, coding_model=CODING_MODEL,
                  enabled_risks=None, coding_strictness=CODING_STRICTNESS):
    """
    コードブック策定（または既存コードブックの読み込み）を行い、指定件数だけコーディング・集計する。
    itemsは{'id','text','fa_no','attrs'}を持つ回答のリスト（アップロードUI側で構築済み）。
    sample_size=0: コーディングしない（策定のみ）／ None: 全件 ／ それ以外: min(sample_size, 全件数)件
    coding_model・enabled_risksはこの分析（result）に紐づけて保存し、続きをコーディングする際も
    同じ設定を使う（分析の途中でモデルやリスクチェック対象が混在しないようにするため）。
    """
    enabled_risks = enabled_risks or []
    reset_token_usage()
    client = make_client('Anthropic', api_key)
    all_items = list(items)
    random.shuffle(all_items)

    codebook, codes = _build_codebook_and_codes(
        client, all_items, max_codes, q_name, data_context, progress_bar, status_text,
        codebook_mode, existing_codebook
    )
    if not codebook:
        return None

    total_items = len(all_items)
    n = total_items if sample_size is None else min(sample_size, total_items)

    results = _code_items(client, q_name, codes, all_items[:n], progress_bar, status_text,
                           coding_model=coding_model, enabled_risks=enabled_risks,
                           coding_strictness=coding_strictness)

    status_text.markdown('**集計中...**')
    gt, sent_counts, unassigned, risk_counts, answer_type_counts = aggregate_results(
        codes, results, n, risk_keys=enabled_risks)
    progress_bar.progress(1.0)
    status_text.markdown('**✅ 完了！**')

    usage = get_token_usage()
    return {
        'codebook':           codebook,
        'codes':              codes,
        'items':              all_items,
        'results':            results,
        'coded_count':        n,
        'total_items':        total_items,
        'gt':                 gt,
        'sent':               sent_counts,
        'unassigned':         unassigned,
        'risk_counts':        risk_counts,
        'enabled_risks':      enabled_risks,
        'answer_type_counts': answer_type_counts,
        'q_name':             q_name,
        'usage':              usage,
        'coding_model':       coding_model,
        'coding_strictness':  coding_strictness,
    }


def _start_coding_job(kind, api_key, q_name, codes, items, target_count, history_id,
                       coding_model, enabled_risks, coding_strictness,
                       reset_usage=True, prior_usage=None):
    """
    バッチ単位で1回のスクリプト再実行ごとに1バッチだけ処理する、中断可能なコーディングジョブを開始する。
    kind='initial'：分析開始ボタン（コードブック策定は既に同期実行済み、reset_usage=Falseで呼ぶ＝
    策定分のトークンを引き継ぐ）。kind='recode'：「現在のコードブックでコーディングする」ボタン
    （reset_usage=Trueで呼び、このジョブ単体のトークン消費をprior_usageに加算する。4.6節参照）。
    """
    if reset_usage:
        reset_token_usage()
    st.session_state.coding_job = {
        'kind':              kind,
        'api_key':           api_key,
        'q_name':            q_name,
        'codes':             codes,
        'items':             items[:target_count],
        'batch_index':       0,
        'results':           [],
        'coding_model':      coding_model,
        'enabled_risks':     enabled_risks or [],
        'coding_strictness': coding_strictness,
        'history_id':        history_id,
        'prior_usage':       prior_usage,
        'stop_requested':    False,
    }


def _finalize_coding_job(job):
    """ジョブの結果（中断時は途中までの分）を集計し、対応する作業履歴エントリに反映する"""
    n = len(job['results'])
    gt, sent_counts, unassigned, risk_counts, answer_type_counts = aggregate_results(
        job['codes'], job['results'], n, risk_keys=job['enabled_risks'])

    usage = get_token_usage()
    if job.get('prior_usage'):
        prior = job['prior_usage']
        usage = {
            'input':          usage.get('input', 0)          + prior.get('input', 0),
            'output':         usage.get('output', 0)         + prior.get('output', 0),
            'cache_read':     usage.get('cache_read', 0)     + prior.get('cache_read', 0),
            'cache_creation': usage.get('cache_creation', 0) + prior.get('cache_creation', 0),
            'cost_jpy':       usage.get('cost_jpy', 0)       + prior.get('cost_jpy', 0),
        }

    finalized = {
        'results':            job['results'],
        'coded_count':        n,
        'gt':                 gt,
        'sent':               sent_counts,
        'unassigned':         unassigned,
        'risk_counts':        risk_counts,
        'enabled_risks':      job['enabled_risks'],
        'answer_type_counts': answer_type_counts,
        'usage':              usage,
        'coding_model':       job['coding_model'],
        'coding_strictness':  job['coding_strictness'],
    }
    for h in st.session_state.history:
        if h['id'] == job['history_id']:
            h['result'].update(finalized)
            break


def _render_coding_job():
    """
    進行中のコーディングジョブを1バッチだけ進め、進捗と「中断する」ボタンを表示してrerunする。
    Streamlitは1回のスクリプト実行が終わるまで新しい操作を処理しないため、for文で全バッチを
    一気に処理すると「中断する」ボタンを押しても反応しない。1バッチ処理するたびにst.rerun()で
    スクリプトを終了・再開することで、バッチの合間（実際のAPI呼び出しで数秒かかる）に中断ボタンの
    クリックを処理できるようにしている。
    """
    job     = st.session_state.coding_job
    items   = job['items']
    batches = [items[i:i+CODING_BATCH_SIZE] for i in range(0, len(items), CODING_BATCH_SIZE)]
    total_batches = len(batches)
    processed     = min(job['batch_index'] * CODING_BATCH_SIZE, len(items))

    st.markdown('#### コーディング中...')
    st.progress(job['batch_index'] / total_batches if total_batches else 1.0)
    st.markdown(f'**{processed}/{len(items)}件処理済み**')
    if st.button('⏹ 中断する（それまでの結果を保存）', width='stretch'):
        job['stop_requested'] = True

    if job['stop_requested'] or job['batch_index'] >= total_batches:
        _finalize_coding_job(job)
        st.session_state.coding_job = None
        st.rerun()
        return

    client = make_client('Anthropic', job['api_key'])
    batch  = batches[job['batch_index']]
    res = _code_one_batch(client, job['q_name'], job['codes'], batch,
                           job['coding_model'], job['enabled_risks'], job['coding_strictness'])
    job['results'].extend(res)
    job['batch_index'] += 1
    st.rerun()


def _start_diagnostic(api_key, q_name, codes, items, target_count, history_id, coding_model):
    """
    精度診断ジョブを開始する。同じ対象件数を「標準」→「厳密」の順でテストコーディングし、
    両者でコードの選択が入れ替わった頻度の高いペアを検出、コードブック見直し案を作成する。
    リスクチェックはコード判定の混同とは無関係なため、診断のコーディングには含めない
    （コスト削減・診断結果の焦点を絞るため）。
    """
    st.session_state.diagnostic_job = {
        'phase':        'standard',
        'api_key':      api_key,
        'q_name':       q_name,
        'codes':        codes,
        'items':        items[:target_count],
        'coding_model': coding_model,
        'batch_index':  0,
        'std_results':  [],
        'strict_results': [],
        'history_id':   history_id,
    }


def _diagnose_confused_pairs(std_results, strict_results, items, top_n=10):
    """
    標準/厳密の2種類のコーディング結果を回答IDで突き合わせ、同じ回答に対して選ばれたコードが
    入れ替わっていた（標準側にしかないコードと厳密側にしかないコードが両方存在する）頻度の高い
    コードペアを検出する。頻度上位top_n件と、ペアごとの回答例（最大3件）を返す。
    """
    item_text = {x['id']: x['text'] for x in items}
    std_by_id    = {r['id']: set(r.get('codes', [])) for r in std_results}
    strict_by_id = {r['id']: set(r.get('codes', [])) for r in strict_results}

    pair_counter = {}
    pair_examples = {}
    for rid, scodes in std_by_id.items():
        tcodes = strict_by_id.get(rid, set())
        only_s = scodes - tcodes
        only_t = tcodes - scodes
        if only_s and only_t:
            for cs in only_s:
                for ct in only_t:
                    key = tuple(sorted([cs, ct]))
                    pair_counter[key] = pair_counter.get(key, 0) + 1
                    examples = pair_examples.setdefault(key, [])
                    if len(examples) < 3:
                        examples.append(item_text.get(rid, ''))

    top_pairs = sorted(pair_counter.items(), key=lambda kv: -kv[1])[:top_n]
    return top_pairs, pair_examples


def _build_diagnostic_instruction(top_pairs, pair_examples, codes_by_id):
    """検出した混同ペア＋回答例から、llm_edit_codebookに渡すコードブック見直し指示文を組み立てる"""
    lines = [
        '以下は、同じ回答セットを「標準」判定基準と「厳密」判定基準の2通りでコーディングした結果、'
        '選ばれるコードが入れ替わっていた頻度の高いペアです。各ペアについて、実際に入れ替わった回答例を示します。',
        '',
    ]
    for (c1, c2), cnt in top_pairs:
        def1 = codes_by_id.get(c1, {})
        def2 = codes_by_id.get(c2, {})
        lines.append(f'■ {c1}「{def1.get("code_name", "")}」 <-> {c2}「{def2.get("code_name", "")}」（{cnt}件で選択が入れ替わり）')
        lines.append(f'  {c1}の定義: {def1.get("definition", "")}')
        lines.append(f'  {c2}の定義: {def2.get("definition", "")}')
        for ex in pair_examples.get((c1, c2), []):
            lines.append(f'  回答例: {ex}')
        lines.append('')
    lines.append(
        '各ペアについて、意味が本質的に重なっており区別する価値が薄いと判断できる場合は1つのコードに統合してください'
        '（統合後はどちらか一方のコードID・コード名を残し、もう一方は削除する）。'
        '区別する価値がある場合は、上記の回答例から「どちらに該当するか」を見分けられるよう、'
        '判断基準を具体的に含めて定義文を書き直してください。'
        '上記に挙げていないコードは変更しないでください。'
    )
    return '\n'.join(lines)


def _render_diagnostic_job():
    """
    進行中の精度診断ジョブを進める。標準→厳密の順でバッチ単位のテストコーディングを行い
    （_render_coding_jobと同じくバッチごとにst.rerun()して中断ボタンに反応できるようにする）、
    両方完了したら混同ペアを分析し、llm_edit_codebookでコードブック見直し案を作成して
    result['pending_edit']にセットする（既存の編集案プレビュー・確定/キャンセルの仕組みをそのまま使う）。
    中断すると診断全体を中止し、それまでの部分結果は使わない（見直し案は作成しない）。
    """
    job = st.session_state.diagnostic_job

    if job['phase'] in ('standard', 'strict'):
        items = job['items']
        batches = [items[i:i+CODING_BATCH_SIZE] for i in range(0, len(items), CODING_BATCH_SIZE)]
        total_batches = len(batches)
        phase_label = '標準' if job['phase'] == 'standard' else '厳密'
        processed = min(job['batch_index'] * CODING_BATCH_SIZE, len(items))

        st.markdown('#### 🎯 精度診断中')
        st.progress(job['batch_index'] / total_batches if total_batches else 1.0)
        st.markdown(f'**{phase_label}でテストコーディング中...（{processed}/{len(items)}件）**')
        if st.button('⏹ 中断する（診断を中止）', width='stretch'):
            st.session_state.diagnostic_job = None
            st.rerun()
            return

        if job['batch_index'] >= total_batches:
            if job['phase'] == 'standard':
                job['phase'] = 'strict'
                job['batch_index'] = 0
            else:
                job['phase'] = 'analyzing'
            st.rerun()
            return

        client = make_client('Anthropic', job['api_key'])
        batch  = batches[job['batch_index']]
        res = _code_one_batch(client, job['q_name'], job['codes'], batch,
                               job['coding_model'], [], job['phase'])
        results_key = 'std_results' if job['phase'] == 'standard' else 'strict_results'
        job[results_key].extend(res)
        job['batch_index'] += 1
        st.rerun()
        return

    # phase == 'analyzing'
    st.markdown('#### 🎯 精度診断中')
    st.markdown('**結果を分析し、コードブック見直し案を作成中...**')

    result = None
    for h in st.session_state.history:
        if h['id'] == job['history_id']:
            result = h['result']
            break

    if result is not None:
        top_pairs, pair_examples = _diagnose_confused_pairs(
            job['std_results'], job['strict_results'], job['items'], top_n=10)
        if top_pairs:
            client = make_client('Anthropic', job['api_key'])
            codes_by_id = {c['code_id']: c for c in job['codes']}
            instruction = _build_diagnostic_instruction(top_pairs, pair_examples, codes_by_id)
            proposed = llm_edit_codebook(client, result['codebook'], instruction, job['q_name'])
            if proposed and proposed.get('categories'):
                result['pending_edit'] = {
                    'instruction': f'🎯 精度診断による見直し提案（混同ペア{len(top_pairs)}件を検出）',
                    'codebook': proposed,
                }
            else:
                result['diagnostic_message'] = '見直し案の作成に失敗しました。再度お試しください。'
        else:
            result['diagnostic_message'] = '標準・厳密の間でコードの選択が入れ替わったペアは見つかりませんでした（コードブックは良好な状態です）。'

    st.session_state.diagnostic_job = None
    st.rerun()


def _render_basic_table_tab(result):
    """
    「基本集計表」タブ：GT集計表からコードを選び（最大2つ、比較用）、該当する文を一覧表示し、
    文をひとつ選ぶと元の回答原文（該当箇所を強調）を表示する、ドリルダウン形式の画面。
    ユーザー提供のレイアウト見本（開発指示用素材/基本集計表.xlsx）に基づく。2026-08-16追加。
    """
    gt      = result.get('gt', [])
    codes   = result.get('codes', [])
    items   = result.get('items', [])
    results = result.get('results', [])

    if not gt or not codes:
        st.info('コーディング結果がまだありません。「🏠 ホーム」タブでコーディングを実行してください。')
        return

    item_map = {it['id']: it for it in items}
    code_by_id = {c['code_id']: c for c in codes}

    # 順A（カテゴリ出現率順→コード出現率順）でコードを並べる。ホームタブのグラフ・Excelと
    # 同じ並び順・同じ判定材料（_category_color_map）を使うことで、表示が食い違わないようにする。
    # 配色（cat_color_full）も同じ関数から取り、GT集計の■■・コード見出しバッジに使う。
    cat_order, cat_color_full = _category_color_map(gt, codes)
    cat_rank     = {cid: i for i, cid in enumerate(cat_order)}
    gt_count_map = {g['code_id']: g['count'] for g in gt}
    gt_pct_map   = {g['code_id']: g['pct'] for g in gt}
    codes_sorted = sorted(
        codes,
        key=lambda c: (cat_rank.get(c['cat_id'], len(cat_order)), -gt_count_map.get(c['code_id'], 0))
    )

    def _on_code_check(code_id):
        """コード選択は最大2つ（比較用）。3つ目を選ぶと最も古く選んだものが自動的に外れる。"""
        key = f'basic_code_{code_id}'
        order = st.session_state.setdefault('basic_code_order', [])
        if st.session_state.get(key):
            if code_id not in order:
                order.append(code_id)
            while len(order) > 2:
                oldest = order.pop(0)
                st.session_state[f'basic_code_{oldest}'] = False
        elif code_id in order:
            order.remove(code_id)

    def _on_sentence_check(sel_key):
        """文選択は常に1件のみ（ドリルダウンが目的のため）。別の文を選ぶと前の選択が自動的に外れる。"""
        prev = st.session_state.get('basic_selected_sentence')
        if st.session_state.get(sel_key):
            if prev and prev != sel_key:
                st.session_state[prev] = False
            st.session_state['basic_selected_sentence'] = sel_key
        elif prev == sel_key:
            st.session_state['basic_selected_sentence'] = None

    col_left, col_right = st.columns([1, 1.4])

    with col_left:
        st.markdown('##### GT集計（コードを選択、最大2つ）')
        prev_cat = None
        for c in codes_sorted:
            if c['cat_id'] != prev_cat:
                cat_color = cat_color_full.get(c['cat_id'], '808080')
                st.markdown(
                    f"<span style='color:#{cat_color};'>■■</span> <strong>{c['cat_name']}</strong>",
                    unsafe_allow_html=True,
                )
                prev_cat = c['cat_id']
            cnt = gt_count_map.get(c['code_id'], 0)
            pct = gt_pct_map.get(c['code_id'], 0.0)
            key = f"basic_code_{c['code_id']}"
            st.session_state.setdefault(key, False)
            st.checkbox(
                f"{c['code_name']}（{pct}%・{cnt}件）",
                key=key, on_change=_on_code_check, args=(c['code_id'],),
            )

    selected_code_ids = st.session_state.get('basic_code_order', [])

    # コード該当文リストの1文分の表示（本文＋属性を1つのmarkdown要素にまとめて行間を詰める）
    def _render_sentence_row(cid, res, it, s, is_partial):
        nonlocal selected_display
        attr_str = '、'.join(f'{k}: {v}' for k, v in it.get('attrs', {}).items() if v)
        text = s.get('text', '')
        line = text
        if attr_str:
            # 半角スペースはHTML上で連続すると詰められてしまうため、崩れずに2文字分の
            # インデントを保てる全角スペースを使う。
            line += f"  \n<span style='font-size:0.78em; color:#888;'>　　{attr_str}</span>"

        # 原文参照（✓欄）の有無に関わらず、常に同じ2カラム構成にしてインデントを揃える
        # （✓欄が無い行だけ左端に寄って見えるという指摘への対応）。✓欄が不要な行は
        # 左カラムに何も置かず空白のまま残す。
        c1, c2 = st.columns([0.06, 0.94], gap=None, vertical_alignment='top')
        if is_partial:
            sel_key = f"basic_sent_{cid}_{res['id']}_{s.get('idx')}"
            st.session_state.setdefault(sel_key, False)
            with c1:
                st.checkbox('', key=sel_key, on_change=_on_sentence_check, args=(sel_key,),
                            label_visibility='collapsed')
            with c2:
                st.markdown(line, unsafe_allow_html=True)
            if st.session_state.get(sel_key):
                selected_display = (it['text'], text)
        else:
            # 文が回答全文と完全に一致する場合は、原文を別途表示する意味が無いため✓欄自体を
            # 出さない（左カラムは空のままでインデントだけ揃える）。
            with c2:
                st.markdown(line, unsafe_allow_html=True)

    with col_right:
        st.markdown('##### 回答原文')
        origin_box = st.container(border=True)

        st.markdown('##### コード該当文リスト')
        if not selected_code_ids:
            st.caption('左のGT集計表でコードを選択してください（最大2つ、比較用）。')

        selected_display = None  # (回答原文全文, 選択した文テキスト)
        # 2コード選択時は、それぞれ独立した固定高さ・スクロール可能な窓に分けて上下に表示する
        # （比較のために2コードまで選べるが、片方の該当文が多いともう片方が下に押し流されて
        # 比較しにくいという指摘への対応）。
        for cid in selected_code_ids:
            code = code_by_id.get(cid)
            if not code:
                continue
            # コード見出しは窓の外（左上）に置き、窓内をスクロールしても常に見える位置に固定する。
            # 配色はGT集計のカテゴリー見出しと同じ「カテゴリーカラーの■■」接頭に統一する。
            cat_color = cat_color_full.get(code['cat_id'], '808080')
            st.markdown(
                f"<span style='color:#{cat_color};'>■■</span> <strong>{code['code_name']}</strong>",
                unsafe_allow_html=True,
            )
            box = st.container(height=340, border=True, gap=None)
            with box:
                found_any = False
                for res in results:
                    it = item_map.get(res.get('id'))
                    if not it:
                        continue
                    sentences = res.get('sentences') or []
                    matched = [s for s in sentences if cid in s.get('codes', [])]
                    if not matched:
                        continue
                    found_any = True
                    if len(matched) > 1:
                        # 同一原文内に同じコードの文が複数ある場合、文ごとに分けて表示すると
                        # 同じ回答者の発言が別々の行に分断されて見えてしまう（ユーザー指摘：
                        # 2026-08-17）。原文全体をまとめて表示する案も一度試したが、原文の
                        # 表示率が上がりすぎるという指摘を受け、自由回答一覧の`_excerpt_for_code`
                        # と同じ方針（該当する文だけを「／」で連結）に統一した。
                        joined = '／'.join(s.get('text', '') for s in matched if s.get('text'))
                        merged = {'text': joined, 'idx': matched[0].get('idx')}
                        is_partial = joined.strip() != it['text'].strip()
                        _render_sentence_row(cid, res, it, merged, is_partial)
                    else:
                        s = matched[0]
                        is_partial = s.get('text', '').strip() != it['text'].strip()
                        _render_sentence_row(cid, res, it, s, is_partial)
                if not found_any:
                    st.caption('該当する文がありません。')

        with origin_box:
            if selected_display:
                full_text, matched = selected_display
                # 複数文が「／」で連結されている場合（同一原文内の同コード文をまとめた行）は、
                # 連結文字列そのままでは原文中に一致しないため、文ごとに分けて個別に強調する。
                shown = full_text
                for part in matched.split('／'):
                    if part and part in shown:
                        shown = shown.replace(part, f'**:orange[{part}]**')
                st.markdown(shown)
            else:
                st.caption('コード該当文リストで部分一致の文（✓欄がある文）を選択すると、'
                           'ここに回答原文（該当箇所を強調）が表示されます。')

    # ── 統合候補・分割候補（コードブック編集タブへ送る前の一時登録、2026-08-19追加） ──
    # 実際にLLMが統合・分割を提案する処理は未実装（今後の設計課題）。ここでは候補を集める
    # ところまでを担う。候補はresult側（プロジェクトごと）に保持し、st.session_state.historyの
    # 該当エントリへ都度書き戻す（他の編集系機能と同じパターン）。
    def _save_basic_table_result():
        for h in st.session_state.history:
            if h['id'] == st.session_state.active_history_id:
                h['result'] = result
                break

    merge_candidates = result.setdefault('merge_candidates', [])
    split_candidates = result.setdefault('split_candidates', [])

    def _claimed_code_ids():
        """統合候補・分割候補のどちらかに既に登録されているコードIDの集合（2026-08-19、
        「片方だけ一致すれば重複登録できてしまう」不具合の指摘を受けて追加）。
        1つのコードは、統合・分割を問わず同時に1つの候補にしか使えない
        （同じコードに対して矛盾する編集案―統合と分割の両方など―が並行して
        作られないようにするための制約。ユーザー判断：「集合的に捉えて現在の
        コーディング結果の総件数に変化がないように」）。"""
        claimed = set()
        for c in merge_candidates:
            claimed.update(c['code_ids'])
        for c in split_candidates:
            claimed.add(c['code_id'])
        return claimed

    st.divider()
    reg_c1, reg_c2 = st.columns(2)
    with reg_c1:
        claimed = _claimed_code_ids()
        merge_disabled = (
            len(selected_code_ids) != 2
            or any(cid in claimed for cid in selected_code_ids)
        )
        if st.button('🔗 統合候補', width='stretch', disabled=merge_disabled,
                     help='選択中の2コードを「統合候補リスト」に一時登録します（コード2つを選択している時のみ。'
                          'いずれかのコードが既に統合・分割の候補に使われている場合は選べません）。'):
            names = [code_by_id[cid]['code_name'] for cid in selected_code_ids if cid in code_by_id]
            # origin_version：登録した時点のコードブックのバージョン（edit_logの件数）を記録する。
            # コードブック編集タブ側で、候補作成後にコードブックが変わっていないかの判定に使う
            # （2026-08-19、コードブック編集タブの設計に合わせて追加）。
            merge_candidates.append({
                'code_ids': list(selected_code_ids), 'code_names': names,
                'origin_version': len(result.get('edit_log', [])),
            })
            _save_basic_table_result()
            st.rerun()
    with reg_c2:
        claimed = _claimed_code_ids()
        split_target = selected_code_ids[0] if len(selected_code_ids) == 1 else None
        split_disabled = split_target is None or split_target in claimed
        if st.button('✂️ 分割候補', width='stretch', disabled=split_disabled,
                     help='選択中の1コードを「分割候補リスト」に一時登録します（コード1つを選択している時のみ。'
                          'そのコードが既に統合・分割の候補に使われている場合は選べません）。'):
            split_candidates.append({
                'code_id': split_target,
                'code_name': code_by_id.get(split_target, {}).get('code_name', split_target),
                'origin_version': len(result.get('edit_log', [])),
            })
            _save_basic_table_result()
            st.rerun()

    st.markdown('##### 🔗 統合候補リスト')
    if not merge_candidates:
        st.caption('まだ登録された統合候補はありません。')
    else:
        for i, cand in enumerate(merge_candidates):
            mc1, mc2 = st.columns([5, 1])
            mc1.markdown('　＋　'.join(cand['code_names']))
            if mc2.button('解除', key=f'merge_cand_release_{i}'):
                merge_candidates.pop(i)
                _save_basic_table_result()
                st.rerun()
        if st.button('📤 統合候補リストをコードブック編集タブに送る', width='stretch', key='send_merge_candidates'):
            edit_queue_merge = result.setdefault('edit_queue_merge', [])
            edit_queue_merge.extend(merge_candidates)
            result['merge_candidates'] = []
            _save_basic_table_result()
            st.rerun()

    st.markdown('##### ✂️ 分割候補リスト')
    if not split_candidates:
        st.caption('まだ登録された分割候補はありません。')
    else:
        for i, cand in enumerate(split_candidates):
            sc1, sc2 = st.columns([5, 1])
            sc1.markdown(cand['code_name'])
            if sc2.button('解除', key=f'split_cand_release_{i}'):
                split_candidates.pop(i)
                _save_basic_table_result()
                st.rerun()
        if st.button('📤 分割候補リストをコードブック編集タブに送る', width='stretch', key='send_split_candidates'):
            edit_queue_split = result.setdefault('edit_queue_split', [])
            edit_queue_split.extend(split_candidates)
            result['split_candidates'] = []
            _save_basic_table_result()
            st.rerun()


def _render_codebook_edit_tab(result, api_key, q_name, diagnostic_size):
    """
    「コードブック編集」タブ（2026-08-19新設）。コードブックの編集に関わる機能を1箇所に
    集約する：①精度診断（標準/厳密比較による見直し提案）②自由チャット編集③基本集計表タブ
    から送られた統合候補・分割候補の処理。①②はコードブック全体を対象とするAI編集案を
    共通のpending_editの仕組み（プレビュー＋確定/キャンセルの2段階確認）で扱う。
    ③は対象コード単位の処理のため独自の仕組みを使う：候補ごとに対象コードの現在の定義を
    提示し、AIに統合後・分割後のコード案（編集可能）を1件だけ作らせ、候補ごとの
    「採用する」（対象コードを削除し新コードを追加）／「中止する」（提案を破棄）で
    即時に確定する（2026-08-19、pending_editの共通プレビューを経由しない専用フローに変更）。
    ①②の編集後は、効果を確かめるために作業メニューの「🧮 コーディング」で再コーディング
    する運用を想定している（このタブ自体はコーディングを行わない）。
    """

    def _save_edit_tab_result():
        for h in st.session_state.history:
            if h['id'] == st.session_state.active_history_id:
                h['result'] = result
                break

    codebook_version = len(result.get('edit_log', []))

    # ── 現在のコードブック（常時表示、表示専用） ──────────────────────
    gt_by_code_current = {g['code_id']: {'count': g['count'], 'pct': g['pct']} for g in result.get('gt', [])}
    st.markdown('#### 📐 現在のコードブック')
    render_codebook_structure(result['codebook'], gt_by_code=gt_by_code_current, key='codebook_edit_current')
    st.divider()

    # ── 編集案のプレビュー（①②③共通、確定・キャンセルの2段階確認） ──────
    pending_edit = result.get('pending_edit')
    if pending_edit:
        st.info(f"📝 編集案：「{pending_edit['instruction']}」（内容を確認して確定してください）")

        removed, added, changed = _diff_codebook(result['codebook'], pending_edit['codebook'])
        if removed or added or changed:
            with st.expander(
                f'🔍 変更点の詳細（削除{len(removed)}・追加{len(added)}・定義や名称の変更{len(changed)}）',
                expanded=True,
            ):
                for c in removed:
                    st.markdown(f"- 🗑️ **削除**：{c.get('code_id')}「{c.get('code_name')}」（他のコードへ統合された可能性があります）")
                for c in added:
                    st.markdown(f"- ➕ **追加**：{c.get('code_id')}「{c.get('code_name')}」")
                for o, n in changed:
                    st.markdown(f"- ✏️ **変更**：{o.get('code_id')}「{o.get('code_name')}」→「{n.get('code_name')}」")
                    st.caption(f"　旧定義：{o.get('definition', '')}")
                    st.caption(f"　新定義：{n.get('definition', '')}")
        else:
            st.caption('※ コード構成に変更はありませんでした（キーワードなど、表に出づらい細部のみの調整である可能性があります）。')

        render_codebook_structure(pending_edit['codebook'], gt_by_code=gt_by_code_current, key='codebook_edit_pending')
        pc1, pc2 = st.columns(2)
        with pc1:
            if st.button('✅ この内容で確定する', type='primary', width='stretch'):
                edit_log = result.setdefault(
                    'edit_log', [{'instruction': '（初期状態）', 'codebook': result['codebook']}]
                )
                edit_log.append({'instruction': pending_edit['instruction'], 'codebook': pending_edit['codebook']})
                result['codebook'] = pending_edit['codebook']
                result['codes'] = [
                    {**c, 'cat_id': cat['cat_id'], 'cat_name': cat['cat_name']}
                    for cat in pending_edit['codebook'].get('categories', [])
                    for c in cat.get('codes', [])
                ]
                # ③（統合候補・分割候補）発の提案だった場合は、確定と同時に該当候補をキューから
                # 取り除く（採用済みなので）。①②発の場合はcandidate_refが無いため何もしない。
                cand_ref = pending_edit.get('candidate_ref')
                if cand_ref:
                    queue_key = 'edit_queue_merge' if cand_ref['list'] == 'merge' else 'edit_queue_split'
                    queue = result.get(queue_key, [])
                    if 0 <= cand_ref['index'] < len(queue):
                        queue.pop(cand_ref['index'])
                del result['pending_edit']
                _save_edit_tab_result()
                st.rerun()
        with pc2:
            if st.button('❌ キャンセル', width='stretch'):
                del result['pending_edit']
                _save_edit_tab_result()
                st.rerun()
        st.divider()

    edit_blocked = pending_edit is not None  # ①②③とも、提案が確定/キャンセルされるまでは新規に作れない

    # ── ① 精度診断（標準・厳密で試しコーディングし、コードの混同を検出して見直し案を作る） ──
    st.markdown('#### ① 精度診断')
    diag_message = result.pop('diagnostic_message', None)
    if diag_message:
        st.info(diag_message)
    if edit_blocked:
        st.caption('※ 上の編集案を確定またはキャンセルしてから精度診断を実行してください。')
    else:
        total_items_for_diag = result.get('total_items', 0)
        diag_target = min(diagnostic_size, total_items_for_diag)
        if st.button(f'🎯 精度診断を実行（{diag_target}件を標準・厳密の両方でテスト）', width='stretch'):
            _start_diagnostic(
                api_key, result.get('q_name', q_name), result['codes'], result['items'],
                diag_target, st.session_state.active_history_id,
                result.get('coding_model', CODING_MODEL),
            )
            st.rerun()
        st.caption('※ 同じ回答を標準・厳密の両方でテストコーディングし、判定が割れやすいコードペアを検出して、'
                   'コードブックの見直し案（統合または定義の書き分け）を自動作成します。'
                   '見直し案は編集案と同じ仕組みで表示され、確定するまでコードブックには反映されません。'
                   '通常のコーディングとは別に2回分のテストコーディング＋見直し案作成のAPIコストがかかります。'
                   '編集を確定したら、効果を確かめるために作業メニューの「🧮 コーディング」で再コーディングしてください。')
    st.divider()

    # ── ② 自由チャット編集（指示文を入力してAIに編集案を作らせる） ──────
    st.markdown('#### ② 自由チャット編集')
    if edit_blocked:
        st.caption('※ 上の編集案を確定またはキャンセルしてからお試しください。')
    else:
        st.caption('統合・改名・再定義・コードの追加や削除など、今あるコードブックの整理を指示できます。'
                   '生データは読み直さないため、新しい観点でのコード発見はできません'
                   '（新しい観点が必要な場合は「🏠 ホーム」の作業メニューでコードブックを作り直してください）。')
        with st.form('edit_instruction_form', clear_on_submit=True):
            instruction = st.text_input(
                '編集の指示を入力', label_visibility='collapsed',
                placeholder='例：AとBのコードを統合して／「対応の速さ」というコードを追加して定義は〜'
            )
            submitted = st.form_submit_button('編集案を作成する', width='stretch')

        if submitted and instruction:
            with st.spinner('編集案を作成中...'):
                client   = make_client('Anthropic', api_key)
                proposed = llm_edit_codebook(client, result['codebook'], instruction, result.get('q_name', q_name))
            if proposed:
                result['pending_edit'] = {'instruction': instruction, 'codebook': proposed}
                _save_edit_tab_result()
                st.rerun()
            else:
                reason = get_last_error() or '原因不明（AIから有効なコードブック構造が返されませんでした）'
                st.error(f'編集案の作成に失敗しました。再度お試しください。\n\n詳細: {reason}')
        st.caption('編集を確定したら、効果を確かめるために作業メニューの「🧮 コーディング」で再コーディングしてください。')
    st.divider()

    # ── ③ 統合候補・分割候補の処理（基本集計表タブから送られたコードにAIが定義案を提案） ──
    st.markdown('#### ③ 統合候補・分割候補の処理')
    edit_queue_merge     = result.get('edit_queue_merge', [])
    edit_queue_split     = result.get('edit_queue_split', [])
    pending_cb_changes   = result.get('pending_codebook_changes', [])

    if not edit_queue_merge and not edit_queue_split and not pending_cb_changes:
        st.info('「📋 基本集計表」タブの「🔗 統合候補」「✂️ 分割候補」から送られたコードが'
                'ここに表示されます。')
        return

    st.caption('「採用する」を押した時点ではコードブックはまだ書き換わりません。統合・分割の作業が'
               '終わったら、下部の「📝 コードブックを書き換える」でまとめて反映してください'
               '（2026-08-19、仕様変更：以前は採用と同時に即時反映していたが、そのたびに他の未処理の'
               '候補が使用不可になり作業が進めにくかったための変更）。'
               '候補を作った時点からコードブックが変わっている（①②で編集済み、リビルド済み、または'
               '別の候補を書き換え済み）場合、その候補は使用できません（コードの前提が変わってしまう'
               'ため）。')

    code_by_id_current = {c['code_id']: c for c in result['codes']}

    st.markdown('##### 🔗 統合候補')
    if not edit_queue_merge:
        st.caption('まだありません。')
    else:
        for i, cand in enumerate(edit_queue_merge):
            is_valid = cand.get('origin_version') == codebook_version
            with st.container(border=True):
                st.markdown('　＋　'.join(cand['code_names']))
                if not is_valid:
                    st.caption('⚠️ 使用不可（登録後にコードブックが変更されています）')
                else:
                    code_a = code_by_id_current.get(cand['code_ids'][0])
                    code_b = code_by_id_current.get(cand['code_ids'][1])
                    if not code_a or not code_b:
                        st.caption('⚠️ 対象コードが見つかりません（既に削除された可能性があります）。')
                    else:
                        st.markdown(f"**{code_a['code_name']}**")
                        st.caption(code_a.get('definition') or '（定義なし）')
                        st.markdown(f"**{code_b['code_name']}**")
                        st.caption(code_b.get('definition') or '（定義なし）')

                        if edit_blocked:
                            st.caption('※ 上の編集案（①②）を確定/キャンセル後に')
                        elif cand.get('proposal') is None:
                            if st.button('🤖 統合案を作成', key=f'propose_merge_{i}', width='stretch'):
                                with st.spinner('統合案を作成中...'):
                                    client = make_client('Anthropic', api_key)
                                    proposed = llm_propose_merge(client, code_a, code_b, result.get('q_name', q_name))
                                if proposed:
                                    cand['proposal'] = proposed
                                    _save_edit_tab_result()
                                    st.rerun()
                                else:
                                    reason = get_last_error() or '原因不明（AIから有効な提案が返されませんでした）'
                                    st.error(f'統合案の作成に失敗しました。再度お試しください。\n\n詳細: {reason}')
                        else:
                            st.markdown('**統合案**（内容は編集できます）')
                            new_name = st.text_input(
                                '統合後のコード名', value=cand['proposal'].get('code_name', ''),
                                key=f'merge_name_{i}',
                            )
                            new_def = st.text_area(
                                '統合後の定義', value=cand['proposal'].get('definition', ''),
                                key=f'merge_def_{i}', height=100,
                            )
                            ac1, ac2 = st.columns(2)
                            with ac1:
                                if st.button('✅ 採用する', type='primary', key=f'adopt_merge_{i}', width='stretch'):
                                    target_cat_id = code_a.get('cat_id')
                                    new_code_name = new_name.strip() or cand['proposal'].get('code_name', '')
                                    staged = result.setdefault('pending_codebook_changes', [])
                                    staged.append({
                                        'kind': 'merge',
                                        'remove_ids': list(cand['code_ids']),
                                        'target_cat_id': target_cat_id,
                                        'new_codes': [{
                                            'code_name':  new_code_name,
                                            'definition': new_def.strip(),
                                            'keywords':   cand['proposal'].get('keywords', []),
                                        }],
                                        'label': f"「{code_a['code_name']}」と「{code_b['code_name']}」を統合して"
                                                 f"「{new_code_name}」を作成（統合候補の採用）",
                                        'source_names': [code_a['code_name'], code_b['code_name']],
                                    })
                                    edit_queue_merge.pop(i)
                                    st.session_state.pop(f'merge_name_{i}', None)
                                    st.session_state.pop(f'merge_def_{i}', None)
                                    _save_edit_tab_result()
                                    st.rerun()
                            with ac2:
                                if st.button('❌ 中止する', key=f'cancel_merge_{i}', width='stretch'):
                                    cand['proposal'] = None
                                    st.session_state.pop(f'merge_name_{i}', None)
                                    st.session_state.pop(f'merge_def_{i}', None)
                                    _save_edit_tab_result()
                                    st.rerun()
                if st.button('候補から削除', key=f'remove_merge_{i}'):
                    edit_queue_merge.pop(i)
                    _save_edit_tab_result()
                    st.rerun()

    st.markdown('##### ✂️ 分割候補')
    if not edit_queue_split:
        st.caption('まだありません。')
    else:
        for i, cand in enumerate(edit_queue_split):
            is_valid = cand.get('origin_version') == codebook_version
            with st.container(border=True):
                st.markdown(cand['code_name'])
                if not is_valid:
                    st.caption('⚠️ 使用不可（登録後にコードブックが変更されています）')
                else:
                    code_x = code_by_id_current.get(cand['code_id'])
                    if not code_x:
                        st.caption('⚠️ 対象コードが見つかりません（既に削除された可能性があります）。')
                    else:
                        st.markdown(f"**{code_x['code_name']}**")
                        st.caption(code_x.get('definition') or '（定義なし）')

                        if edit_blocked:
                            st.caption('※ 上の編集案（①②）を確定/キャンセル後に')
                        elif cand.get('proposals') is None:
                            if st.button('🤖 分割案を作成', key=f'propose_split_{i}', width='stretch'):
                                with st.spinner('分割案を作成中...'):
                                    client = make_client('Anthropic', api_key)
                                    proposed = llm_propose_split(client, code_x, result.get('q_name', q_name))
                                if proposed and len(proposed.get('codes', [])) == 2:
                                    cand['proposals'] = proposed['codes']
                                    _save_edit_tab_result()
                                    st.rerun()
                                else:
                                    reason = get_last_error() or '原因不明（AIから有効な提案が返されませんでした）'
                                    st.error(f'分割案の作成に失敗しました。再度お試しください。\n\n詳細: {reason}')
                        else:
                            st.markdown('**分割案**（内容は編集できます）')
                            new_names, new_defs = [], []
                            for k, prop in enumerate(cand['proposals']):
                                st.markdown(f'分割案{k + 1}')
                                new_names.append(st.text_input(
                                    f'分割案{k + 1}のコード名', value=prop.get('code_name', ''),
                                    key=f'split_name_{i}_{k}', label_visibility='collapsed',
                                ))
                                new_defs.append(st.text_area(
                                    f'分割案{k + 1}の定義', value=prop.get('definition', ''),
                                    key=f'split_def_{i}_{k}', height=80, label_visibility='collapsed',
                                ))
                            ac1, ac2 = st.columns(2)
                            with ac1:
                                if st.button('✅ 採用する', type='primary', key=f'adopt_split_{i}', width='stretch'):
                                    target_cat_id = code_x.get('cat_id')
                                    new_codes = [
                                        {
                                            'code_name':  new_names[k].strip() or cand['proposals'][k].get('code_name', ''),
                                            'definition': new_defs[k].strip(),
                                            'keywords':   cand['proposals'][k].get('keywords', []),
                                        }
                                        for k in range(2)
                                    ]
                                    staged = result.setdefault('pending_codebook_changes', [])
                                    staged.append({
                                        'kind': 'split',
                                        'remove_ids': [cand['code_id']],
                                        'target_cat_id': target_cat_id,
                                        'new_codes': new_codes,
                                        'label': f"「{code_x['code_name']}」を「{new_codes[0]['code_name']}」と"
                                                 f"「{new_codes[1]['code_name']}」に分割（分割候補の採用）",
                                        'source_names': [code_x['code_name']],
                                    })
                                    edit_queue_split.pop(i)
                                    for k in range(2):
                                        st.session_state.pop(f'split_name_{i}_{k}', None)
                                        st.session_state.pop(f'split_def_{i}_{k}', None)
                                    _save_edit_tab_result()
                                    st.rerun()
                            with ac2:
                                if st.button('❌ 中止する', key=f'cancel_split_{i}', width='stretch'):
                                    cand['proposals'] = None
                                    for k in range(2):
                                        st.session_state.pop(f'split_name_{i}_{k}', None)
                                        st.session_state.pop(f'split_def_{i}_{k}', None)
                                    _save_edit_tab_result()
                                    st.rerun()
                if st.button('候補から削除', key=f'remove_split_{i}'):
                    edit_queue_split.pop(i)
                    _save_edit_tab_result()
                    st.rerun()

    st.divider()
    st.markdown('##### 📝 コードブックへの反映待ち')
    if not pending_cb_changes:
        st.caption('「採用する」で作った変更がここに並びます。統合・分割の作業がひと通り終わったら、'
                   '下の「コードブックを書き換える」でまとめて反映してください。')
    else:
        for j, chg in enumerate(pending_cb_changes):
            with st.container(border=True):
                if chg['kind'] == 'merge':
                    new_name = chg['new_codes'][0]['code_name']
                    st.markdown(f"🔗 統合：{chg['source_names'][0]}　＋　{chg['source_names'][1]}　→　**{new_name}**")
                else:
                    new_names = '、'.join(nc['code_name'] for nc in chg['new_codes'])
                    st.markdown(f"✂️ 分割：{chg['source_names'][0]}　→　**{new_names}**")
                if st.button('取り消す', key=f'undo_staged_{j}'):
                    pending_cb_changes.pop(j)
                    _save_edit_tab_result()
                    st.rerun()

        if edit_blocked:
            st.caption('※ 上の編集案（①②）を確定/キャンセル後に書き換えられます。')
        elif st.button('📝 コードブックを書き換える', type='primary', width='stretch', key='commit_staged_changes'):
            _, dupes = _apply_staged_changes(result, pending_cb_changes)
            if dupes:
                st.error('コード名が重複しています（既存のコード、または今回追加する別のコードと重複）：'
                          + '、'.join(dupes) + '　該当する提案のコード名を修正するか、取り消してから'
                          '再度お試しください。')
            else:
                result['pending_codebook_changes'] = []
                _save_edit_tab_result()
                st.success('✅ コードブックを書き換えました。')
                st.rerun()


def _render_cross_tab_tab(result):
    """
    「クロス集計」タブ：属性（年代・職位など、Excel取り込み時に列指定した属性）を1つ選び、
    コード×属性値のクロス集計表（実数・出現率%）と横棒グラフを表示する。2026-08-18追加。
    ％は既存のGT集計と同じ考え方（その属性値グループの有効回答数に対する出現率）で計算する。
    """
    coded_count = result.get('coded_count', 0)
    items   = result.get('items', [])[:coded_count]
    results = result.get('results', [])
    codes   = result.get('codes', [])
    gt      = result.get('gt', [])

    if not codes or not gt or coded_count == 0:
        st.info('コーディング結果がまだありません。「🏠 ホーム」タブでコーディングを実行してください。')
        return

    # 属性キー一覧（自由回答一覧シートと同じ、最初に登場した順）
    attr_keys = []
    for it in items:
        for k in it.get('attrs', {}):
            if k not in attr_keys:
                attr_keys.append(k)

    if not attr_keys:
        st.info('属性列が設定されていません。「📊 Excelファイル（ID・属性列あり）」でアップロードする際に'
                '属性として使う列を選択すると、ここでクロス集計ができます。')
        return

    attr_key = st.selectbox('クロスする属性を選択', attr_keys, key='cross_tab_attr_select')

    if st.button('📊 クロス集計を開始', key='cross_tab_start_btn'):
        st.session_state.cross_tab_active_attr = attr_key

    active_attr = st.session_state.get('cross_tab_active_attr')
    if not active_attr or active_attr not in attr_keys:
        return

    result_map = {r['id']: r for r in results}

    # 属性値ごとの母数（有効回答数）を、初出順で確定する
    attr_values, group_total = [], {}
    for it in items:
        v = str(it.get('attrs', {}).get(active_attr, '') or '（未回答）')
        if v not in group_total:
            group_total[v] = 0
            attr_values.append(v)
        group_total[v] += 1

    code_counts = {}
    for it in items:
        res = result_map.get(it['id'])
        if not res:
            continue
        v = str(it.get('attrs', {}).get(active_attr, '') or '（未回答）')
        for cid in res.get('codes', []):
            code_counts[(cid, v)] = code_counts.get((cid, v), 0) + 1

    # 順A（カテゴリ出現率順→コード出現率順）でコードを並べる。ホーム・基本集計表と同じ判定材料。
    cat_order, cat_color_full = _category_color_map(gt, codes)
    cat_rank     = {cid: i for i, cid in enumerate(cat_order)}
    gt_count_map = {g['code_id']: g['count'] for g in gt}
    codes_sorted = sorted(
        codes,
        key=lambda c: (cat_rank.get(c['cat_id'], len(cat_order)), -gt_count_map.get(c['code_id'], 0))
    )

    # 属性値ごとのグロス（g）＝そのグループでの該当コード数の延べ合計（複数コード該当時はnより
    # 大きくなる。nは人数、gは延べタグ数、というアンケート集計の一般的な区別）。
    gross_by_value = {
        v: sum(code_counts.get((c['code_id'], v), 0) for c in codes_sorted)
        for v in attr_values
    }

    st.divider()
    st.markdown(f'#### 📋 クロス集計表（{active_attr} × コード）')
    st.caption('※ ％は各属性値の有効回答数（n）に対する出現率です。グロス(g)は該当コード数の'
               '延べ合計（複数コード該当時はnより大きくなります）：' +
               '、'.join(f'{v}（n={group_total[v]}・g={gross_by_value[v]}）' for v in attr_values))

    import pandas as pd

    n_row = {'カテゴリー': '', 'コード': 'n（基数・人数）'}
    g_row = {'カテゴリー': '', 'コード': 'グロス(g)（延べタグ数）'}
    for v in attr_values:
        n_row[f'{v}・実数'] = group_total[v]
        n_row[f'{v}・％']   = ''
        g_row[f'{v}・実数'] = gross_by_value[v]
        g_row[f'{v}・％']   = ''

    table_rows  = [n_row, g_row]
    row_cat_ids = [None, None]
    for c in codes_sorted:
        row = {'カテゴリー': c['cat_name'], 'コード': c['code_name']}
        for v in attr_values:
            cnt = code_counts.get((c['code_id'], v), 0)
            total_v = group_total[v]
            row[f'{v}・実数']  = cnt
            row[f'{v}・％']    = round(cnt / total_v * 100, 1) if total_v else 0.0
        table_rows.append(row)
        row_cat_ids.append(c['cat_id'])

    df_table = pd.DataFrame(table_rows)
    cat_col_idx  = df_table.columns.get_loc('カテゴリー')
    code_col_idx = df_table.columns.get_loc('コード')

    def _style_cat_col(row):
        styles = [''] * len(row)
        cat_id = row_cat_ids[row.name]
        if cat_id is not None:
            # ホームの「コード別GT集計」グラフの棒と同じ色（_category_color_map）をそのまま使う。
            styles[cat_col_idx] = f'background-color: #{cat_color_full.get(cat_id, "808080")}; color: #FFFFFF; font-weight: bold;'
        else:
            styles[cat_col_idx]  = 'font-weight: bold; background-color: rgba(128,128,128,0.15);'
            styles[code_col_idx] = 'font-weight: bold; background-color: rgba(128,128,128,0.15);'
        return styles

    st.dataframe(df_table.style.apply(_style_cat_col, axis=1), width='stretch', hide_index=True)
    st.caption('※ カテゴリー欄の背景色は、ホームの「コード別GT集計」グラフの棒と同じ配色です。')

    st.divider()
    st.markdown('#### 📈 クロス集計グラフ')

    import plotly.express as px
    plot_rows = []
    for c in codes_sorted:
        for v in attr_values:
            cnt = code_counts.get((c['code_id'], v), 0)
            total_v = group_total[v]
            plot_rows.append({
                'コード名': c['code_name'],
                active_attr: v,
                '件数': cnt,
                '出現率(%)': round(cnt / total_v * 100, 1) if total_v else 0.0,
            })
    df_plot = pd.DataFrame(plot_rows)
    code_order_top_to_bottom = [c['code_name'] for c in codes_sorted]
    fig = px.bar(
        df_plot,
        y='コード名',
        x='出現率(%)',
        color=active_attr,
        orientation='h',
        barmode='group',
        category_orders={'コード名': list(reversed(code_order_top_to_bottom))},
        labels={'出現率(%)': '出現率(%)', 'コード名': ''},
        hover_data=['件数'],
    )
    # コード名（y軸ラベル）の先頭にカテゴリー色の█を付け、表側と同じ配色でカテゴリーを示す。
    # （■よりも視認性の高い太字ブロック文字█を使う。2026-08-19、ユーザー要望で■から変更）
    cat_id_by_name = {c['code_name']: c['cat_id'] for c in codes_sorted}
    tick_names = list(reversed(code_order_top_to_bottom))
    tick_text = [
        f'<span style="color:#{cat_color_full.get(cat_id_by_name.get(name), "808080")}">█</span> {name}'
        for name in tick_names
    ]
    # 凡例（下記）の文字サイズをここと明示的に揃えるため、既定値任せにせず固定値を指定する
    # （2026-08-19、ユーザー要望：「凡例の文字サイズをグラフ内のコード名のサイズと同じに」）。
    TICK_FONT_SIZE = 12
    fig.update_yaxes(tickmode='array', tickvals=tick_names, ticktext=tick_text,
                      tickfont=dict(size=TICK_FONT_SIZE))
    fig.update_layout(
        height=max(420, 42 * len(codes_sorted)),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
        margin=dict(l=10),
    )
    st.plotly_chart(fig, width='stretch')
    st.caption('※ コード名の左の█は、下のカテゴリー凡例に対応する色です。')

    # カテゴリーの色凡例（グラフの色は属性値ごとの比較に使っているため、カテゴリーの色は
    # 別途、コード名接頭の█と対応するこの凡例で示す。2026-08-19、ユーザー要望で追加）。
    legend_html = '　'.join(
        f'<span style="color:#{cat_color_full.get(cid, "808080")}">█</span> {name}'
        for cid, name in {c['cat_id']: c['cat_name'] for c in codes_sorted}.items()
    )
    st.markdown(
        f'<div style="text-align:center; font-size:{TICK_FONT_SIZE}px;">{legend_html}</div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════
# バブル図（放射状：カテゴリー→コード→代表引用文、2026-08-19追加）
# ══════════════════════════════════════════════════

def _xml_escape(s):
    """SVGのtext要素に埋め込む文字列をエスケープする（&/</>/"）。"""
    return str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


def _svg_wrap_lines(text, max_chars):
    """日本語は単語区切りが無いため、文字数ベースで単純に折り返す。"""
    text = str(text)
    lines = [text[i:i + max_chars] for i in range(0, len(text), max_chars)]
    return lines or ['']


def _svg_text_block(x, y, text, max_chars, font_size, fill, anchor='middle', font_weight='normal', max_lines=2):
    """複数行に折り返したcenter/start/end基準のSVG<text>要素を返す（縦方向も中央寄せ）。"""
    lines = _svg_wrap_lines(text, max_chars)
    truncated = len(lines) > max_lines
    lines = lines[:max_lines]
    if truncated and lines:
        lines[-1] = (lines[-1][:-1] if len(lines[-1]) >= max_chars else lines[-1]) + '…'
    n = len(lines)
    start_y = y - (n - 1) * font_size * 0.6
    tspans = ''.join(
        f'<tspan x="{x:.1f}" dy="{0 if i == 0 else font_size * 1.2:.1f}">{_xml_escape(line)}</tspan>'
        for i, line in enumerate(lines)
    )
    return (f'<text x="{x:.1f}" y="{start_y:.1f}" font-size="{font_size}" fill="#{fill}" '
            f'text-anchor="{anchor}" font-weight="{font_weight}" font-family="Meiryo, sans-serif">'
            f'{tspans}</text>')


def _bubble_size_tier(value, max_value):
    """該当数から大中小3段階のサイズ区分を返す（完全比例ではなく大まかな差で十分という
    ユーザー要望に対応。上位1/3を「大」、中間1/3を「中」、それ以外を「小」とする）。"""
    if max_value <= 0:
        return 'small'
    ratio = value / max_value
    if ratio >= 0.66:
        return 'large'
    if ratio >= 0.33:
        return 'medium'
    return 'small'


def _representative_quotes(results, code_id, max_quotes=3, max_chars=20):
    """指定コードが付与された文を最大max_quotes件集める（バブル図でコードバブルの周辺に
    表示する代表引用文）。同じ回答から複数文は採用しない（多様な回答を見せるため、
    1回答につき最初に該当した1文のみ）。長い文はmax_chars文字で切り詰める。"""
    quotes = []
    for res in results:
        if len(quotes) >= max_quotes:
            break
        for s in res.get('sentences') or []:
            if code_id in s.get('codes', []) and s.get('text'):
                text = s['text']
                if len(text) > max_chars:
                    text = text[:max_chars] + '…'
                quotes.append(text)
                break
    return quotes


# カテゴリーの文字サイズは、cat_font（基準値＝中サイズ扱い）にこの倍率を掛けて大中小を
# 自動で作る（ユーザー要望：「カテゴリーの文字サイズはバブルサイズに合わせて最初から
# 大中小になっていてほしい」、2026-08-19追加）。
_CAT_FONT_TIER_MULT = {'large': 1.25, 'medium': 1.0, 'small': 0.8}
BUBBLE_CANVAS_SIZE = 2000  # SVGキャンバスの一辺（px）。画面表示スケール調整でも参照する。

BUBBLE_PARAM_DEFAULTS = {
    # カテゴリー：全体の中心からの距離／バブルサイズ倍率／文字サイズ（中サイズの基準値）／折り返し文字数
    # 2026-08-19、ユーザーが実機で調整した値を初期値として採用（400/1.70/24/8）。
    'cat_dist':  400, 'cat_size': 1.7, 'cat_font': 24, 'cat_chars': 8,
    # コード：所属カテゴリーの中心からの距離（基本値）／バブルサイズ倍率／文字サイズ／折り返し文字数
    # 同上（210/1.50/13/7）。
    'code_dist': 210, 'code_size': 1.5, 'code_font': 13, 'code_chars': 7,
    # 文（代表引用文）：コードバブルからの距離（基本値）／文字サイズ／切り詰め文字数／表示件数上限／
    # コード→文の引き線の線幅／引き線の濃度（不透明度）
    # （ラベルの枠を廃止したため、枠の幅パラメータは持たない。2026-08-19）
    'quote_dist': 26, 'quote_font': 10, 'quote_chars': 20, 'quote_max_count': 3,
    'quote_line_width': 1.0, 'quote_line_opacity': 0.4,
}


def _build_bubble_svg(gt, codes, results, cat_color_full, selected_cat_ids, q_name, params=None):
    """
    バブル図のSVG文字列を生成する：円周上に配置したカテゴリーバブル（中央から放射状）を
    起点に、各カテゴリーの周りにそのカテゴリーのコードバブルを扇状に配置し、細い線で結ぶ。
    さらに各コードバブルの外側に、該当する代表的な原文（最大3件）を短く切り詰めて添える。
    大画面・ポスター表示を想定し、ベクター形式（SVG）で生成する——ラスター画像と違い
    どれだけ拡大しても文字が滲まないため（ユーザー要望：「拡大すると読めるといい」）。
    バブルの大きさは該当数に応じた大中小3段階（_bubble_size_tier）を基本に、paramsの
    サイズ倍率でさらに拡大縮小する（完全比例ではなく大中小の差＋ユーザー調整の組み合わせ）。
    色はホームの「コード別GT集計」グラフと同じ_category_color_map配色を使い、コード
    バブルは同じ色の淡色版（_lighten_hex、Excel出力と共通）でカテゴリーとの親子関係を示す。

    params：カテゴリー・コード・文（引用文）ごとに、バブルサイズ・親バブル中心からの距離・
    文字サイズ・文字数を調整できるパラメータ辞書（BUBBLE_PARAM_DEFAULTSの形。省略時は既定値）。
    2026-08-19、ユーザー要望「カテゴリー、コード、文ごとに、バブルサイズ、バブルの中心
    （上位バブルの中心）からの距離、文字サイズ、文字数」を調整可能にするために追加。
    コードの距離・文の距離は、密集を避けるための自動調整（コード数に応じた開き角の拡大、
    文の段積みオフセット）をベース値に上乗せする形は維持しており、パラメータは「基準値」を
    動かすものとして扱う（密集回避の自動調整自体は無効化しない）。
    """
    p = {**BUBBLE_PARAM_DEFAULTS, **(params or {})}

    W = H = BUBBLE_CANVAS_SIZE
    CX, CY = W / 2, H / 2
    R_CAT = p['cat_dist']
    CAT_R  = {k: v * p['cat_size'] for k, v in {'large': 95, 'medium': 75, 'small': 55}.items()}
    CODE_R = {k: v * p['code_size'] for k, v in {'large': 40, 'medium': 32, 'small': 24}.items()}

    gt_count_by_code = {g['code_id']: g['count'] for g in gt}
    cat_total = {}
    for g in gt:
        if g['cat_id'] in selected_cat_ids:
            cat_total[g['cat_id']] = cat_total.get(g['cat_id'], 0) + g['count']

    cats, seen_cat = [], set()
    for c in codes:
        if c['cat_id'] in selected_cat_ids and c['cat_id'] not in seen_cat:
            cats.append({'cat_id': c['cat_id'], 'cat_name': c['cat_name']})
            seen_cat.add(c['cat_id'])
    cats.sort(key=lambda c: -cat_total.get(c['cat_id'], 0))  # 順Aと同じ、出現数の多い順

    codes_by_cat = {}
    for c in codes:
        if c['cat_id'] in selected_cat_ids:
            codes_by_cat.setdefault(c['cat_id'], []).append(c)
    for cid in codes_by_cat:
        codes_by_cat[cid].sort(key=lambda c: -gt_count_by_code.get(c['code_id'], 0))

    max_cat_total  = max(cat_total.values(), default=0)
    max_code_count = max(
        (gt_count_by_code.get(c['code_id'], 0) for c in codes if c['cat_id'] in selected_cat_ids),
        default=0,
    )

    n_cat = len(cats)
    line_svgs, code_svgs, quote_svgs, cat_svgs = [], [], [], []

    for i, cat in enumerate(cats):
        theta = (2 * math.pi * i / n_cat - math.pi / 2) if n_cat else 0.0
        cat_cx = CX + R_CAT * math.cos(theta)
        cat_cy = CY + R_CAT * math.sin(theta)
        color = cat_color_full.get(cat['cat_id'], '808080')
        cat_tier = _bubble_size_tier(cat_total.get(cat['cat_id'], 0), max_cat_total)
        cat_r = CAT_R[cat_tier]
        cat_font_size = round(p['cat_font'] * _CAT_FONT_TIER_MULT[cat_tier])

        cat_svgs.append(f'<circle cx="{cat_cx:.1f}" cy="{cat_cy:.1f}" r="{cat_r:.1f}" '
                         f'fill="#{color}" stroke="#FFFFFF" stroke-width="3"/>')
        cat_svgs.append(_svg_text_block(cat_cx, cat_cy, cat['cat_name'], p['cat_chars'], cat_font_size, 'FFFFFF',
                                         font_weight='bold', max_lines=3))

        cat_codes = codes_by_cat.get(cat['cat_id'], [])
        n_codes = len(cat_codes)
        if n_codes == 0:
            continue
        spread_deg = min(160, max(50, n_codes * 20))
        code_radius_offset = p['code_dist'] + min(90, n_codes * 6)
        code_fill = _lighten_hex(color, 0.6)

        for j, code in enumerate(cat_codes):
            angle_offset_deg = 0.0 if n_codes == 1 else (-spread_deg / 2 + spread_deg * j / (n_codes - 1))
            code_theta = theta + math.radians(angle_offset_deg)
            code_cx = cat_cx + code_radius_offset * math.cos(code_theta)
            code_cy = cat_cy + code_radius_offset * math.sin(code_theta)

            code_count = gt_count_by_code.get(code['code_id'], 0)
            code_r = CODE_R[_bubble_size_tier(code_count, max_code_count)]

            line_svgs.append(f'<line x1="{cat_cx:.1f}" y1="{cat_cy:.1f}" x2="{code_cx:.1f}" y2="{code_cy:.1f}" '
                              f'stroke="#{color}" stroke-width="1.5" opacity="0.55"/>')
            code_svgs.append(f'<circle cx="{code_cx:.1f}" cy="{code_cy:.1f}" r="{code_r:.1f}" '
                              f'fill="#{code_fill}" stroke="#{color}" stroke-width="2"/>')
            code_svgs.append(_svg_text_block(code_cx, code_cy, code['code_name'], p['code_chars'], p['code_font'],
                                              '333333', font_weight='600', max_lines=3))

            quotes = _representative_quotes(results, code['code_id'], max_quotes=p['quote_max_count'],
                                             max_chars=p['quote_chars'])
            if quotes:
                n_quotes = len(quotes)
                # 引用文どうしの重なりを減らすための工夫（ユーザー提案の「斜め」「放射状」＋実データでの
                # 見た目を受けた追加調整、2026-08-19）：
                # ①同じコードの複数の引用文を、コードの向きを中心に扇状に広げる（単純に1本の直線上に
                #  並べるより隣の文とぶつかりにくい）。扇の開き角は、文字数が多いほど実際に必要な
                #  横幅も広がるため、最長の引用文の推定文字幅から動的に計算する（固定26度だけでは
                #  長い引用文がひしめくケースで重なりが残っていたため）。
                # ②隣り合うコード（j mod 3）で引用文までの距離を3段階でずらし、隣接コードの引用文が
                #  同じ「輪」に重なって並ぶのを避ける（当初は2段階だったが、コード数が多いカテゴリー
                #  では3段階の方が余裕ができる）。
                max_q_chars = max(len(q) for q in quotes)
                est_text_width = max_q_chars * p['quote_font'] * 0.95
                avg_r = code_r + p['quote_dist']
                min_fan_deg = math.degrees(est_text_width / max(avg_r, 50)) * 1.3
                fan_spread_deg = min(80, max(26, min_fan_deg))
                stagger = (j % 3) * (p['quote_dist'] * 1.3)
                for k, q in enumerate(quotes):
                    fan_offset_deg = 0.0 if n_quotes == 1 else (-fan_spread_deg / 2 + fan_spread_deg * k / (n_quotes - 1))
                    quote_theta = code_theta + math.radians(fan_offset_deg)
                    qdx, qdy = math.cos(quote_theta), math.sin(quote_theta)
                    anchor = 'start' if qdx >= 0.15 else ('end' if qdx <= -0.15 else 'middle')
                    q_r = code_r + p['quote_dist'] + stagger + k * (p['quote_dist'] * 1.3)
                    qx, qy = code_cx + qdx * q_r, code_cy + qdy * q_r
                    # コード→文の引き線（黒、線幅・濃度は調整可能。2026-08-19追加）。カテゴリー→コードの
                    # 線と同じline_svgsに入れることで、コード・カテゴリーの各バブルの下に描画され、
                    # バブルの縁からきれいに伸びているように見える（バブルの上に線が乗らない）。
                    line_svgs.append(f'<line x1="{code_cx:.1f}" y1="{code_cy:.1f}" x2="{qx:.1f}" y2="{qy:.1f}" '
                                      f'stroke="#000000" stroke-width="{p["quote_line_width"]}" '
                                      f'opacity="{p["quote_line_opacity"]}"/>')
                    quote_svgs.append(_svg_text_block(qx, qy, q, p['quote_chars'] + 2, p['quote_font'], '555555',
                                                       anchor=anchor, max_lines=1))

    title = f'{q_name}：コード分布バブル図' if q_name else 'コード分布バブル図'
    svg = (
        # width/heightを明示しないと、st.html側のoverflow:autoなdivが高さ0に潰れて
        # 何も表示されなくなる（実機テストで発見。SVGは既定でwidth/height省略時「100%」
        # 扱いになり、高さ未確定の親要素の中では0に解決されるため）。
        f'<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 {W} {H}" width="{W}" height="{H}" '
        f'font-family="Meiryo, sans-serif">'
        f'<rect x="0" y="0" width="{W}" height="{H}" fill="#FFFFFF"/>'
        # タイトルは左上・レギュラー体・サイズは元の2/3（34→約23）に変更（2026-08-19、ユーザー要望）
        + _svg_text_block(40, 45, title, 60, 23, '333333', anchor='start', font_weight='normal', max_lines=1)
        + ''.join(line_svgs) + ''.join(code_svgs) + ''.join(quote_svgs) + ''.join(cat_svgs)
        + '</svg>'
    )
    return svg


def _render_bubble_tab(result):
    """
    「🫧 バブル図」タブ：中央から放射状にカテゴリー→コード→代表引用文を配置したバブル図を
    生成する。ポスター・大画面での提示を想定し、ベクター（SVG）で生成・ダウンロードできる。
    2026-08-19追加。カテゴリー数・コード数が多い場合（ユーザー実績：最大10カテゴリー×
    40コード程度を想定）に画面が混み合いすぎる場合は、カテゴリーを複数回に分けて別々の
    バブル図を生成できるよう、対象カテゴリーを選べるようにしている（ユーザー要望：
    「難しい場合は2枚に分けることも考えます」への対応）。
    """
    coded_count = result.get('coded_count', 0)
    gt      = result.get('gt', [])
    codes   = result.get('codes', [])
    results = result.get('results', [])

    if not codes or not gt or coded_count == 0:
        st.info('コーディング結果がまだありません。「🏠 ホーム」タブでコーディングを実行してください。')
        return

    cat_order, cat_color_full = _category_color_map(gt, codes)
    cat_name_by_id = {}
    for c in codes:
        cat_name_by_id.setdefault(c['cat_id'], c['cat_name'])
    cat_options = [cid for cid in cat_order if cid in cat_name_by_id]

    st.caption('中央から放射状に「カテゴリー→コード→代表的な原文」を配置したバブル図を作成します。'
               'ポスターや大画面での提示を想定しており、SVG（ベクター形式）で生成するため'
               '画面上で文字が小さく見えても、拡大すれば滲まずに読めます。')

    selected_cats = st.multiselect(
        '対象カテゴリーを選択（多すぎて見づらい場合は複数回に分けて生成してください）',
        options=cat_options,
        default=cat_options,
        format_func=lambda cid: cat_name_by_id.get(cid, cid),
        key='bubble_selected_cats',
    )

    with st.expander('⚙️ 詳細設定（バブルサイズ・距離・文字サイズ・文字数）'):
        st.caption('カテゴリー・コード・文（代表引用文）それぞれについて、親バブル中心からの'
                   '距離・バブルサイズ・文字サイズ・文字数を調整できます。')
        pc1, pc2, pc3 = st.columns(3)
        with pc1:
            st.markdown('**カテゴリー**')
            cat_dist  = st.slider('全体の中心からの距離', 200, 900, BUBBLE_PARAM_DEFAULTS['cat_dist'],
                                   step=20, key='bubble_cat_dist')
            cat_size  = st.slider('バブルサイズ倍率', 0.5, 2.0, BUBBLE_PARAM_DEFAULTS['cat_size'],
                                   step=0.1, key='bubble_cat_size')
            cat_font  = st.slider('文字サイズ', 12, 40, BUBBLE_PARAM_DEFAULTS['cat_font'],
                                   step=1, key='bubble_cat_font')
            cat_chars = st.slider('文字数（折り返し）', 4, 20, BUBBLE_PARAM_DEFAULTS['cat_chars'],
                                   step=1, key='bubble_cat_chars')
        with pc2:
            st.markdown('**コード**')
            code_dist  = st.slider('カテゴリー中心からの距離', 80, 300, BUBBLE_PARAM_DEFAULTS['code_dist'],
                                    step=10, key='bubble_code_dist')
            code_size  = st.slider('バブルサイズ倍率', 0.5, 2.0, BUBBLE_PARAM_DEFAULTS['code_size'],
                                    step=0.1, key='bubble_code_size')
            code_font  = st.slider('文字サイズ', 8, 24, BUBBLE_PARAM_DEFAULTS['code_font'],
                                    step=1, key='bubble_code_font')
            code_chars = st.slider('文字数（折り返し）', 4, 16, BUBBLE_PARAM_DEFAULTS['code_chars'],
                                    step=1, key='bubble_code_chars')
        with pc3:
            st.markdown('**文（代表引用文）**')
            quote_dist  = st.slider('コードバブルからの距離', 10, 80, BUBBLE_PARAM_DEFAULTS['quote_dist'],
                                     step=2, key='bubble_quote_dist')
            quote_font  = st.slider('文字サイズ', 6, 16, BUBBLE_PARAM_DEFAULTS['quote_font'],
                                     step=1, key='bubble_quote_font')
            quote_chars = st.slider('文字数（切り詰め）', 8, 40, BUBBLE_PARAM_DEFAULTS['quote_chars'],
                                     step=2, key='bubble_quote_chars')
            quote_max_count = st.slider('表示件数の上限（コードごと）', 0, 5,
                                         BUBBLE_PARAM_DEFAULTS['quote_max_count'],
                                         step=1, key='bubble_quote_max_count',
                                         help='重なりが気になる場合、件数を減らすと密集が緩和されます。')
            quote_line_width = st.slider('コード→文の引き線の太さ', 0.5, 4.0,
                                          BUBBLE_PARAM_DEFAULTS['quote_line_width'],
                                          step=0.5, key='bubble_quote_line_width')
            quote_line_opacity = st.slider('コード→文の引き線の濃度', 0.0, 1.0,
                                            BUBBLE_PARAM_DEFAULTS['quote_line_opacity'],
                                            step=0.05, key='bubble_quote_line_opacity',
                                            help='0で非表示、1で真っ黒（不透明度）になります。')
        st.caption('※ コードの「距離」・文の「距離」は基準値です。バブルが密集しすぎないよう、'
                   'コード数や引用文の件数に応じた自動調整が上乗せされます。引用文どうしの重なりは'
                   '自動でも軽減していますが、完全ではありません。気になる場合は「表示件数の上限」を'
                   '減らすか、文字数を短くしてみてください。')

    if st.button('🫧 バブル図を生成', key='bubble_generate_btn'):
        st.session_state.bubble_svg_cats = list(selected_cats)
        st.session_state.bubble_svg_params = {
            'cat_dist': cat_dist, 'cat_size': cat_size, 'cat_font': cat_font, 'cat_chars': cat_chars,
            'code_dist': code_dist, 'code_size': code_size, 'code_font': code_font, 'code_chars': code_chars,
            'quote_dist': quote_dist, 'quote_font': quote_font, 'quote_chars': quote_chars,
            'quote_max_count': quote_max_count,
            'quote_line_width': quote_line_width, 'quote_line_opacity': quote_line_opacity,
        }

    active_cats = st.session_state.get('bubble_svg_cats')
    if not active_cats:
        return

    n_codes_selected = sum(1 for c in codes if c['cat_id'] in active_cats)
    st.caption(f'※ 選択中：{len(active_cats)}カテゴリー・{n_codes_selected}コード')

    active_params = st.session_state.get('bubble_svg_params', BUBBLE_PARAM_DEFAULTS)
    svg = _build_bubble_svg(gt, codes, results, cat_color_full, set(active_cats), result.get('q_name', ''),
                             params=active_params)

    # 画面表示スケール（2026-08-19追加、ユーザー要望：「画面の表示スケールが変えられないか」）。
    # ダウンロード用のsvg（実寸2000x2000、width/height属性つき）はそのまま、画面表示用にだけ
    # width/height属性を置き換えたコピーを作る。SVGはviewBoxを保ったままwidth/heightだけ
    # 変えるとベクターのまま拡大縮小されるため、画質の劣化なくスケールできる。既定45%は、
    # 900pxの表示枠に全体がだいたい収まる大きさ（2000px×45%≒900px）。
    display_scale = st.slider('画面表示の倍率(%)', 20, 150, 45, step=5, key='bubble_display_scale',
                               help='画面上の見た目だけを調整します。ダウンロードするSVGファイルは'
                                    '常に実寸（2000×2000）のままです。')
    display_px = round(BUBBLE_CANVAS_SIZE * display_scale / 100)
    svg_display = svg.replace(
        f'width="{BUBBLE_CANVAS_SIZE}" height="{BUBBLE_CANVAS_SIZE}"',
        f'width="{display_px}" height="{display_px}"',
        1,
    )

    # st.html()はDOMPurifyでサニタイズされ、<svg>要素がまるごと除去されてしまうため
    # （実機テストで発見）、生HTMLをそのまま埋め込めるst.iframe()を使う。表示はこのタブの
    # 自作SVGのみで外部/ユーザー入力コンテンツではないため、st.iframeのraw HTML経路を
    # 使うリスクは無い（quoteテキストは_xml_escapeで別途エスケープ済み）。
    # 画面上は900pxの枠内にスクロール表示する（実サイズは2000x2000あるため、そのまま置くと
    # 縦に長大になりすぎる）。印刷・ポスター用の実寸はダウンロードしたSVGファイルの方を使う。
    st.iframe(
        '<div style="width:100%; height:900px; overflow:auto; box-sizing:border-box; '
        'border:1px solid #E0E0E0; border-radius:8px; padding:8px;">' + svg_display + '</div>',
        height=920,
    )

    st.download_button(
        label='📥 バブル図をダウンロード（SVG）',
        data=svg.encode('utf-8'),
        file_name=f'AfterCoding_bubble_{datetime.now().strftime("%Y%m%d_%H%M")}.svg',
        mime='image/svg+xml',
        width='stretch',
    )
    st.caption('※ SVGファイルはブラウザや多くの画像編集ソフトで開けます。印刷・ポスター用途では'
               'そのまま拡大しても画質が劣化しません。')


# ══════════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════════

# 固定タイトル（👻アフターコーディング支援ツール）は2026-08-19に削除。ブランディングは
# サイドバー（アフターコーディング支援ツール After Coder）に統合済みで、右画面はタブの
# 表示面積を最大化するため、上部の固定見出しは持たない。

# コーディングジョブ・精度診断ジョブが進行中は、進捗＋中断ボタンだけを表示してスクリプトを終了する
# （サイドバー・結果表示など他のUIは中断可否に関わらず操作させない）。
if st.session_state.coding_job:
    _render_coding_job()
    st.stop()
if st.session_state.diagnostic_job:
    _render_diagnostic_job()
    st.stop()

# プロジェクトファイル読み込み直後のプロジェクト名反映。st.session_state['q_name_input']は
# 下のsidebarでウィジェットが生成された後は直接書き換えられない（Streamlitの制約）ため、
# 読み込み処理側は代わりに'_pending_q_name'に値を置いてrerunし、ウィジェット生成より前の
# ここで確定させる（2026-08-19、実機テストで発見・修正）。
if '_pending_q_name' in st.session_state:
    st.session_state['q_name_input'] = st.session_state.pop('_pending_q_name')

# ── サイドバー：設定 ──────────────────────────────
with st.sidebar:
    st.markdown(
        "<div class='sidebar-title-wrap'>"
        "<p class='sidebar-title-jp'>アフターコーディング支援ツール</p>"
        "<p class='sidebar-title-en'>After Coder</p>"
        "</div>",
        unsafe_allow_html=True,
    )
    st.markdown('*KOTONOHA by Marketing Junction*')
    st.caption(f'👤 ログイン中: {st.session_state.username}')
    st.divider()
    st.header('⚙️ 設定')
    api_key = st.text_input(
        'APIキー',
        type='password',
        placeholder='sk-ant-...',
        help='Anthropic ConsoleでAPIキーを取得してください'
    )
    q_name = st.text_input(
        'プロジェクト名',
        placeholder='例：サービスへのご意見・ご感想',
        help='分析する自由回答設問の名前を入力してください（プロジェクトファイル名やレポート見出しにも使われます）',
        key='q_name_input',
    )
    data_context = st.text_area(
        label='分析データの特徴',
        placeholder='例：青森県の小中高校教員を対象とした教育改革に関するアンケートです。回答者は教員・校長・教頭・学校職員です。',
        height=120,
        help='コーディング作業時にAIが参考に使用します。調査内容や目的、調査対象者、質問内容など、データの特徴について記入してください。（記入しないでも作動はしますが記入した方が精度が上がります）'
    )
    max_codes = st.slider(
        'コード数の上限',
        min_value=10, max_value=49, value=30, step=1,
        help='生成するコードの最大数（推奨：20〜35）'
    )

    st.markdown('**📐 コードブック策定方式**')
    codebook_mode_label = st.selectbox(
        'コードブック策定方式',
        CODEBOOK_MODE_LABELS,
        index=CODEBOOK_MODE_LABELS.index('方式C-自動：全件精査（件数自動調整）（推奨）'),
        label_visibility='collapsed',
    )
    codebook_mode = CODEBOOK_MODE_CODE[codebook_mode_label]

    n_texts = st.session_state.texts_count
    if n_texts >= 2000 and codebook_mode in ('A', 'B'):
        st.info(f'📊 {n_texts}件のデータには方式C-自動（またはC1〜C3）を推奨します')
    elif n_texts >= 500 and codebook_mode == 'A':
        st.info(f'📊 {n_texts}件のデータには方式B以上を推奨します')

    existing_codebook_data = None
    if codebook_mode == 'EXISTING':
        existing_file = st.file_uploader(
            'コードブックファイル（CSV または JSON、過去バージョンも可）',
            type=['csv', 'json'],
            help='「コードブック」または「コード一覧集計」の表右上のツールバーからダウンロードしたCSV、'
                 'または以前保存したJSONファイルを指定してください（過去のバージョンをアップロードして編集することもできます）'
        )
        if existing_file:
            try:
                if existing_file.name.lower().endswith('.csv'):
                    existing_codebook_data = parse_codebook_csv(existing_file.read())
                else:
                    existing_codebook_data = json.loads(existing_file.read().decode('utf-8'))
                n_loaded_codes = sum(
                    len(cat.get('codes', [])) for cat in existing_codebook_data.get('categories', [])
                )
                st.success(f'✅ コードブックを読み込みました（{n_loaded_codes}コード）')
            except Exception as e:
                st.error(f'ファイルの読み込みに失敗しました: {e}')
        st.caption('※ 既存のコードブックを使用する場合、「コード数の上限」は適用されません')

    st.markdown('**🧮 コーディング範囲**')
    coding_scope_label = st.selectbox(
        'コーディング範囲',
        CODING_SCOPE_OPTIONS,
        index=0,
        label_visibility='collapsed',
        help='作業メニューの「コードブックの作成とコーディング」「コーディング」を実行する際に、'
             'どこまでコーディングするかを選びます。まず少量でコーディングし、結果を見ながら'
             'コードブックを編集し、納得できたら「全件コーディング」で範囲を広げる使い方を想定しています。'
             '（「コードブックの作成」単体はこの設定を使わず、常にコーディングなしで実行されます）'
    )
    coding_sample_size = CODING_SCOPE_SIZES[coding_scope_label]

    coding_model_label = st.selectbox(
        'コーディングモデル',
        list(CODING_MODEL_OPTIONS.keys()),
        index=0,
        help='コーディング（分類作業）に使うモデルを選びます。Haiku 4.5はSonnet 4.6の約1/3の単価です。'
             '精度・価格を比較したい場合は、同じコードブックのまま方式を変えて2回実行し、'
             '「🕒 最近使用したプロジェクト」で結果とコストを見比べてください（1回の分析中でモデルが混在することはありません）。'
    )
    coding_model = CODING_MODEL_OPTIONS[coding_model_label]

    coding_strictness_label = st.selectbox(
        'コーディング判定の厳密度',
        list(CODING_STRICTNESS_OPTIONS.keys()),
        index=0,
        help='コードを付与する際の判定基準の厳しさです。「厳密」はコードの定義・キーワードに明確に一致する場合のみ付与し、'
             '誤検出（過剰付与）を減らします（非該当が増える可能性があります）。「柔軟」は表現が違っても幅広く付与し、'
             '見落としを減らします（過剰付与が増える可能性があります）。リスクチェック・回答分類の判定基準には影響しません。'
    )
    coding_strictness = CODING_STRICTNESS_OPTIONS[coding_strictness_label]

    diagnostic_max = max(20, min(300, n_texts)) if n_texts else 300
    diagnostic_size = st.slider(
        '🎯 精度診断のテスト件数',
        min_value=20, max_value=diagnostic_max, value=min(100, diagnostic_max), step=10,
        help='精度診断（標準・厳密の2種類でテストコーディングし、コードブックの見直し案を作成する機能）で使う件数です。'
             '通常のコーディング範囲とは別に、少なめの件数で素早く診断する用途を想定しています。'
    )

    st.markdown('**⚠️ リスクチェック（該当有無を検知）**')
    enabled_risks = []
    for opt in RISK_CHECK_OPTIONS:
        if st.checkbox(opt['label'], help=opt['hint'], key=f"risk_{opt['key']}"):
            enabled_risks.append(opt['key'])
    st.caption('「不明」「無回答」はチェック不要（常に判定）')

    st.divider()
    st.markdown('**💾 プロジェクトファイル**')
    st.markdown('**🕒 最近使用したプロジェクト**')
    if st.session_state.history:
        for h in reversed(st.session_state.history[-3:]):
            ts    = h['timestamp'].strftime('%m/%d %H:%M')
            label = h['q_name'] if len(h['q_name']) <= 18 else h['q_name'][:18] + '…'
            active = h['id'] == st.session_state.active_history_id
            if st.button(
                f"{'▶ ' if active else ''}{ts}　{label}",
                key=f"hist_btn_{h['id']}",
                width='stretch',
                type='primary' if active else 'secondary',
            ):
                st.session_state.active_history_id = h['id']
                st.rerun()
    else:
        st.caption('まだプロジェクトがありません')

    active_result_for_save = next(
        (h['result'] for h in st.session_state.history if h['id'] == st.session_state.active_history_id),
        None
    )
    if active_result_for_save:
        project_payload = _build_project_file(active_result_for_save)
        st.download_button(
            '💾 プロジェクトファイルをダウンロード',
            data=json.dumps(project_payload, ensure_ascii=False, indent=2, default=str),
            file_name=f'AfterCoderProject_{datetime.now().strftime("%Y%m%d_%H%M")}.json',
            mime='application/json',
            width='stretch',
        )
    else:
        st.caption('（保存できる分析結果がまだありません）')

    opened_project_file = st.file_uploader(
        '📂 プロジェクトファイルを開く', type=['json'], key='project_file_uploader',
        help='以前ダウンロードしたプロジェクトファイル（JSON）を読み込み、保存時点の状態から再開します。'
    )
    if opened_project_file is not None:
        if st.button('📂 この内容で開く', width='stretch', key='open_project_file_btn'):
            try:
                loaded_payload = json.loads(opened_project_file.read().decode('utf-8'))
                loaded_result  = _load_project_file(loaded_payload)
            except Exception as e:
                st.error(f'プロジェクトファイルの読み込みに失敗しました: {e}')
            else:
                st.session_state.history_counter += 1
                hist_id = st.session_state.history_counter
                st.session_state.history.append({
                    'id':        hist_id,
                    'q_name':    loaded_result.get('q_name', ''),
                    'timestamp': datetime.now(),
                    'result':    loaded_result,
                })
                st.session_state.history = st.session_state.history[-10:]
                st.session_state.active_history_id = hist_id
                # プロジェクト名の入力欄（key='q_name_input'）も読み込んだ内容に合わせておく。
                # ウィジェット生成後にq_name_inputへ直接代入するとStreamlitAPIExceptionになる
                # ため、'_pending_q_name'に置いてrerunし、スクリプト冒頭（ウィジェット生成前）で
                # 反映させる。これをしないと、読み込み直後はプロジェクト名が空欄のままになり、
                # 「コードブックの作成」（RAWデータからのリビルド）がq_name未入力を理由に
                # 押せなくなってしまう（2026-08-18、実機テストで発見・修正）。
                st.session_state['_pending_q_name'] = loaded_result.get('q_name', '')
                st.success('✅ プロジェクトファイルを読み込みました')
                st.rerun()

    st.divider()
    if st.button('🔄 設定をリセット（新しい分析用）', width='stretch',
                 help='アップロード中のExcelデータやリスクチェックのチェック状態など、'
                      '分析条件だけをリセットします。APIキー・作業履歴は保持されます。'
                      '新しいデータで分析を始める前に、古い設定が残っていないか不安な場合に押してください。'):
        st.session_state.xlsx_items = []
        for k in ('xlsx_text_col', 'xlsx_id_col', 'xlsx_fa_col', 'xlsx_attr_cols'):
            st.session_state.pop(k, None)
        for opt in RISK_CHECK_OPTIONS:
            st.session_state.pop(f"risk_{opt['key']}", None)
        st.session_state.texts_count = 0
        st.rerun()

    st.divider()
    if st.button('🚪 ログアウト', width='stretch'):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.rerun()

# ── メインエリア ──────────────────────────────────
# タブを画面最上部に固定表示する（2026-08-19）。作業メニュー・設定確認・結果表示のすべてを
# 🏠ホームタブの中に入れ、プロジェクトが無い状態でもタブバー自体は常に見える状態にする。
tab_home, tab_basic, tab_edit, tab_cross, tab_bubble = st.tabs(
    ['🏠 ホーム', '📋 基本集計表', '🔧 コードブック編集', '📊 クロス集計', '🫧 バブル図']
)

with tab_home:

    with st.expander('📖 使い方・プロジェクトファイルについて'):
        st.markdown('''
1. APIキーとプロジェクト名を入力
2. 分析データの特徴を記入（任意）
3. 作業メニューでファイルをアップロード
4. 作業メニューの実行ボタンをクリック
5. 結果を確認してExcelをダウンロード
        ''')
        st.caption('RAWデータ・コードブック・コーディング結果を1つにまとめて保存し、'
                   '次回はアップロードのやり直しなしで続きから再開できます（サイドバー'
                   '「💾 プロジェクトファイル」）。')

    # 「今アクティブなプロジェクト」の参照。col1（作業メニュー）・col2（設定確認）の両方が使うため、
    # ここで1回だけ引く。col2の実行ボタン（プロジェクトの作成のみ／コードブックの作成／作成と
    # コーディング／コーディング）はいずれも完了後にst.rerun()するため、この直後の判定は常に
    # 「ボタンを押す前」の状態を見ている。本計算（結果表示に使う方）はボタン処理の後に改めて行う。
    _active_result_now = next(
        (h['result'] for h in st.session_state.history if h['id'] == st.session_state.active_history_id),
        None
    )

    col1, col2 = st.columns([1, 1])

    with col1:
        st.markdown('### 📋 新規プロジェクト作成')
        st.caption('注・既存のプロジェクトファイルを使う場合は、サイドバー左下「📂 プロジェクトファイルを開く」から。')

        with st.expander('📁 ファイルアップロード', expanded=(st.session_state.texts_count == 0)):
            input_method = st.radio(
                '入力方法を選択',
                ['📄 テキストファイル', '📋 テキストを直接貼り付け', '📊 Excelファイル（ID・属性列あり）'],
                horizontal=True
            )

            items = []

            if input_method == '📄 テキストファイル':
                uploaded = st.file_uploader(
                    'テキストファイルを選択（1行1回答）',
                    type=['txt'],
                    help='UTF-8形式のテキストファイル。1行に1つの回答を記入してください。'
                )
                if uploaded:
                    content = uploaded.read().decode('utf-8', errors='ignore')
                    texts   = [line.strip() for line in content.splitlines() if line.strip()]
                    items   = _items_from_texts(texts)
                    st.success(f'✅ {len(items)}件の回答を読み込みました')
                    with st.expander('回答プレビュー（先頭5件）'):
                        for i, it in enumerate(items[:5], 1):
                            st.markdown(f'**{i}.** {it["text"]}')

            elif input_method == '📋 テキストを直接貼り付け':
                pasted = st.text_area(
                    'テキストを貼り付け（1行1回答）',
                    height=200,
                    placeholder='例：\n対応が丁寧で安心できた\n待ち時間が長かった\n先生の説明がわかりやすかった',
                    help='1行に1つの回答を入力してください。空行は自動的に除外されます。'
                )
                if pasted:
                    texts = [line.strip() for line in pasted.splitlines() if line.strip()]
                    items = _items_from_texts(texts)
                    st.success(f'✅ {len(items)}件の回答を読み込みました')
                    with st.expander('回答プレビュー（先頭5件）'):
                        for i, it in enumerate(items[:5], 1):
                            st.markdown(f'**{i}.** {it["text"]}')

            else:
                import pandas as pd
                uploaded = st.file_uploader(
                    'Excelファイルを選択',
                    type=['xlsx'],
                    help='自由記述に加えて、回答者ID・FA番号・年代や職位などの属性列を含むExcelファイルを読み込めます。'
                         '属性は集計・コーディングの判断には使われず、レポート表示のためだけに保持されます。'
                )
                if uploaded:
                    df = pd.read_excel(uploaded)
                    st.caption(f'{len(df)}行を読み込みました。列の役割を指定してください。')
                    st.dataframe(df.head(), width='stretch')

                    text_col = st.selectbox('自由記述の列', df.columns, key='xlsx_text_col')
                    id_col   = st.selectbox('回答者IDの列（任意）', ['(なし・自動採番)'] + list(df.columns), key='xlsx_id_col')
                    fa_col   = st.selectbox('FA番号の列（任意）', ['(なし)'] + list(df.columns), key='xlsx_fa_col')
                    attr_cols = st.multiselect(
                        '属性として使う列（複数選択可・任意。年代・職位など）',
                        [c for c in df.columns if c not in {text_col, id_col, fa_col}],
                        key='xlsx_attr_cols'
                    )

                    if st.button('この内容で読み込む'):
                        built = []
                        for i, row in df.iterrows():
                            text = str(row[text_col]).strip() if pd.notna(row[text_col]) else ''
                            rid  = (str(row[id_col]) if id_col != '(なし・自動採番)' and pd.notna(row[id_col])
                                    else f'NO{i+1:03d}')
                            fa   = str(row[fa_col]) if fa_col != '(なし)' and pd.notna(row[fa_col]) else None
                            attrs = {c: row[c] for c in attr_cols}
                            built.append({'id': rid, 'text': text, 'fa_no': fa, 'attrs': attrs})
                        st.session_state.xlsx_items = built

                    items = st.session_state.get('xlsx_items', [])
                    if items:
                        st.success(f'✅ {len(items)}件の回答を読み込みました')
                        with st.expander('回答プレビュー（先頭5件）'):
                            for i, it in enumerate(items[:5], 1):
                                attr_str = '、'.join(f'{k}: {v}' for k, v in it['attrs'].items())
                                st.markdown(f'**{i}.** [{it["id"]}] {it["text"]}' + (f'　（{attr_str}）' if attr_str else ''))

        if st.session_state.texts_count != len(items):
            st.session_state.texts_count = len(items)
            st.rerun()

        st.markdown('##### 📝 コードブック作成時の要望')
        codebook_request = st.text_area(
            'コードブック作成時の要望', label_visibility='collapsed',
            placeholder='例：「対応の速さ」と「説明の分かりやすさ」は別コードにしてください／'
                        '否定的な意見も肯定的な意見と同じ粒度で細分化してください　など',
            height=100, key='codebook_request_input',
            help='コードブックを新規に策定・リビルドする際、この内容を優先的に反映してAIに'
                 '作成させます（下の「分析データの特徴」より優先度の高い、具体的な指示として扱われます）。'
        )

    with col2:
        st.markdown('### 📋 作業メニュー')

        # 「コードブックの作成」系ボタンをリビルドとして使う場合、col1で新たにファイルを
        # アップロードしていなくても、読み込み済みプロジェクトが持つRAWデータ（items）を
        # そのまま使えるようにする。そうしないと、既存プロジェクトを開いただけでは常に
        # items=[]（col1は今回未アップロード）のままとなり、リビルドボタンが押せなくなる
        # （2026-08-18、実機テストで発見・修正）。新規アップロードがあればそちらを優先する。
        rebuild_items = items if items else (_active_result_now.get('items', []) if _active_result_now else [])

        # コードブック作成時の要望（col1）は、既存の「分析データの特徴」（サイドバー、data_context）
        # より優先度の高い指示として、その直前に強調して連結する。data_context自体は変えず、
        # コードブック生成呼び出し（run_pipeline/_build_codebook_and_codes）にだけ渡す専用の
        # 変数にする（2026-08-19追加。要望はコードブック生成のみに使い、コーディング自体には使わない）。
        effective_data_context = (
            (f'【最優先の要望・必ず反映してください】\n{codebook_request.strip()}\n\n' if codebook_request.strip() else '')
            + data_context
        )

        mode_ready         = codebook_mode != 'EXISTING' or existing_codebook_data is not None
        project_only_ready = bool(items and q_name)
        upload_ready        = bool(rebuild_items and q_name and api_key and mode_ready)
        coding_ready         = bool(_active_result_now and _active_result_now.get('codebook') and api_key)

        if st.button('📁 プロジェクトの作成のみ', type='primary', width='stretch', disabled=not project_only_ready,
                     help='コードブック・コーディングは行わず、アップロードした内容と設定だけをプロジェクトとして'
                          '保存します（APIキーは不要）。後でこのプロジェクトを開き、続きから作成できます。'):
            st.session_state.history_counter += 1
            hist_entry = {
                'id':        st.session_state.history_counter,
                'q_name':    q_name,
                'timestamp': datetime.now(),
                'result': {
                    'codebook':           None,
                    'codes':              [],
                    'items':              list(items),
                    'results':            [],
                    'coded_count':        0,
                    'total_items':        len(items),
                    'gt':                 [],
                    'sent':               {'positive': 0, 'negative': 0, 'neutral': 0},
                    'unassigned':         0,
                    'risk_counts':        {},
                    'enabled_risks':      enabled_risks,
                    'answer_type_counts': {},
                    'q_name':             q_name,
                    'usage':              {'input': 0, 'output': 0, 'cache_read': 0, 'cache_creation': 0, 'cost_jpy': 0.0},
                    'coding_model':       coding_model,
                    'coding_strictness':  coding_strictness,
                },
            }
            st.session_state.history.append(hist_entry)
            st.session_state.history = st.session_state.history[-10:]
            st.session_state.active_history_id = hist_entry['id']
            st.rerun()

        st.caption('「コードブックの作成」系ボタンは、既にコードブックがある状態で押すと、'
                   'RAWデータから作り直す「リビルド」になります。')

        if st.button('📐 コードブックの作成', type='primary', width='stretch', disabled=not upload_ready,
                     help='コーディングは行わず、コードブックの生成のみを行います。'):
            progress_bar = st.progress(0)
            status_text  = st.empty()
            with st.spinner('処理中...'):
                result = run_pipeline(
                    api_key, q_name, rebuild_items, max_codes,
                    progress_bar, status_text, effective_data_context, codebook_mode, existing_codebook_data,
                    0, coding_model, enabled_risks, coding_strictness
                )
            if result:
                st.session_state.history_counter += 1
                hist_entry = {
                    'id':        st.session_state.history_counter,
                    'q_name':    q_name,
                    'timestamp': datetime.now(),
                    'result':    result,
                }
                st.session_state.history.append(hist_entry)
                st.session_state.history = st.session_state.history[-10:]
                st.session_state.active_history_id = hist_entry['id']
                # ここでrerunしないと、この下にある「🧮 コーディング」ボタンのcoding_ready判定が
                # このスクリプト実行の先頭で計算した古い_active_result_now（コードブック作成前）
                # のままになり、実際にはコードブックができたのにボタンが無効のままになってしまう
                # （2026-08-19、実機テストで発見・修正）。
                st.rerun()

        if st.button('📐➡️🧮 コードブックの作成とコーディング', type='primary', width='stretch', disabled=not upload_ready,
                     help='コードブックを生成した後、サイドバーの「コーディング範囲」で指定した件数までコーディングします。'):
            reset_token_usage()
            client = make_client('Anthropic', api_key)
            all_items = list(rebuild_items)
            random.shuffle(all_items)

            progress_bar = st.progress(0)
            status_text  = st.empty()
            with st.spinner('コードブックを策定中...'):
                codebook, codes = _build_codebook_and_codes(
                    client, all_items, max_codes, q_name, effective_data_context, progress_bar, status_text,
                    codebook_mode, existing_codebook_data
                )

            if codebook:
                total_items = len(all_items)
                target = total_items if coding_sample_size is None else min(coding_sample_size, total_items)

                st.session_state.history_counter += 1
                hist_id = st.session_state.history_counter
                base_result = {
                    'codebook':           codebook,
                    'codes':              codes,
                    'items':              all_items,
                    'results':            [],
                    'coded_count':        0,
                    'total_items':        total_items,
                    'gt':                 [],
                    'sent':               {'positive': 0, 'negative': 0, 'neutral': 0},
                    'unassigned':         0,
                    'risk_counts':        {},
                    'enabled_risks':      enabled_risks,
                    'answer_type_counts': {},
                    'q_name':             q_name,
                    'usage':              get_token_usage(),
                    'coding_model':       coding_model,
                    'coding_strictness':  coding_strictness,
                }
                st.session_state.history.append({
                    'id':        hist_id,
                    'q_name':    q_name,
                    'timestamp': datetime.now(),
                    'result':    base_result,
                })
                st.session_state.history = st.session_state.history[-10:]
                st.session_state.active_history_id = hist_id

                _start_coding_job('initial', api_key, q_name, codes, all_items, target, hist_id,
                                   coding_model, enabled_risks, coding_strictness, reset_usage=False)
                st.rerun()

        if st.button('🧮 コーディング（現在のコードブックを使用）', type='primary', width='stretch', disabled=not coding_ready,
                     help='今表示中のコードブックはそのままに、サイドバーの「コーディング範囲」で指定した件数を'
                          '最初からコーディングし直します。'):
            _active_total  = _active_result_now.get('total_items', len(_active_result_now.get('items', [])))
            coding_target  = _active_total if coding_sample_size is None else min(coding_sample_size, _active_total)
            _start_coding_job(
                'recode', api_key, _active_result_now.get('q_name', q_name),
                _active_result_now['codes'], _active_result_now['items'],
                coding_target, st.session_state.active_history_id,
                _active_result_now.get('coding_model', CODING_MODEL),
                _active_result_now.get('enabled_risks', []),
                _active_result_now.get('coding_strictness', CODING_STRICTNESS),
                reset_usage=True, prior_usage=_active_result_now.get('usage'),
            )
            st.rerun()
        if not coding_ready:
            st.caption('※ 使用できるコードブックがまだありません。先に「コードブックの作成」を行うか、'
                       'サイドバーからプロジェクトファイルを開いてください。')

        st.divider()
        st.markdown('##### ✅ 設定確認')

        has_project  = bool(items) or _active_result_now is not None
        has_codebook = bool(_active_result_now and _active_result_now.get('codebook'))
        has_coding   = bool(_active_result_now and _active_result_now.get('coded_count', 0) > 0)

        def _status_badge(label, done):
            return f"{'✅' if done else '⬜'} {label}"

        st.markdown(
            f"{_status_badge('プロジェクト登録', has_project)}　"
            f"{_status_badge('コードブック', has_codebook)}　"
            f"{_status_badge('コーディング結果', has_coding)}"
        )
        st.divider()

        # 2026-08-19：以前は items/q_name/api_key が揃うまで案内文だけを表示していたが、
        # サイドバーの設定内容は常に確認できた方が良いという指摘を受け、条件分岐をやめて
        # 「設定状況」として常時表示するようにした（回答数・処理時間目安はitems依存のため
        # アップロード後のみ追加表示する）。
        st.markdown('##### 設定状況')
        st.markdown(f'**プロジェクト名：** {q_name or "（未入力）"}')
        st.markdown(f'**コード数上限：** {max_codes}個')
        st.markdown(f'**策定方式：** {codebook_mode_label}')
        st.markdown(f'**コーディングモデル：** {coding_model_label}')
        risk_labels = [o['label'] for o in RISK_CHECK_OPTIONS if o['key'] in enabled_risks]
        st.markdown(f'**リスクチェック：** {"、".join(risk_labels) if risk_labels else "なし"}')
        st.markdown(f'**APIキー：** {"設定済み ✅" if api_key else "未設定"}')
        if items:
            st.markdown(f'**回答数：** {len(items)}件')
            est_min = max(3, len(items) // 100 * 2)
            st.info(f'⏱️ 処理時間の目安：{est_min}〜{est_min*2}分')

    st.divider()

    # 作業メニューのボタン処理より後で改めて引き直す（同期パス完了直後の反映を保つため。
    # 冒頭の_active_result_nowはボタン処理"前"の状態で、col1/col2の判定にのみ使う）。
    active_result = None
    active_q_name = None
    for h in st.session_state.history:
        if h['id'] == st.session_state.active_history_id:
            active_result = h['result']
            active_q_name = h['q_name']
            break

    if active_result:
        result = active_result
        q_name = active_q_name

        coded_count = result.get('coded_count', 0)
        total_items = result.get('total_items', 0)

        # コスト表示（共通。プロンプトキャッシュの読み込み/書き込み分は呼び出し時点のモデル単価で
        # 都度計算し累積している＝コードブック策定と方式が混在していても正確な合計になる）。
        # 2026-08-19、作業メニュー／設定確認の直下に表示されるよう位置を変更した。
        usage      = result.get('usage', {})
        inp        = usage.get('input', 0)
        out        = usage.get('output', 0)
        cache_read = usage.get('cache_read', 0)
        cache_new  = usage.get('cache_creation', 0)
        cost       = usage.get('cost_jpy', 0.0)
        with st.expander('💰 API使用コスト（参考）'):
            c1, c2, c3 = st.columns(3)
            c1.metric('入力トークン', f'{inp:,}')
            c2.metric('出力トークン', f'{out:,}')
            c3.metric('推定コスト', f'約 ¥{cost:.0f}')
            if cache_read or cache_new:
                c4, c5 = st.columns(2)
                c4.metric('キャッシュ読込（約1/10単価）', f'{cache_read:,}')
                c5.metric('キャッシュ書込（約1.25倍単価）', f'{cache_new:,}')
            st.caption(
                '※ 1USD=150円換算。コードブック策定はSonnet、コーディングはHaikuの料金に基づく概算です。'
                'コーディングはコードブックをプロンプトキャッシュしており、2回目以降のバッチはキャッシュ読込分が割安になります。'
            )

        if not result.get('codebook'):
            st.success('✅ プロジェクトを保存しました！')
        elif coded_count == 0:
            st.success('✅ コードブックの生成が完了しました！')
        elif coded_count < total_items:
            st.success(f'✅ {coded_count}/{total_items}件をコーディングしました！')
        else:
            st.success('✅ 分析が完了しました！')
        # 「この分析のコーディングモデル」キャプションは2026-08-19にホームから削除した
        # （設定確認の「設定状況」で同じ情報を確認できるため）。

        if not result.get('codebook'):
            # 「📁 プロジェクトの作成のみ」で作られたプロジェクト（コードブック未作成）の場合。
            st.info('このプロジェクトはまだコードブックがありません。作業メニューの'
                    '「📐 コードブックの作成」または「📐➡️🧮 コードブックの作成とコーディング」を'
                    '実行してください。')
        else:
            # ── コードブック（GT数値付き、表示専用）──────────────────────
            # 2026-08-19、編集関連の機能（編集案プレビュー・確定/キャンセル・精度診断・
            # 自由チャット編集）は、すべて新設の「🔧 コードブック編集」タブへ移設した。
            # ホームは常に「今のコードブックがどうなっているか」の確認だけを担う。
            gt_by_code_current = {g['code_id']: {'count': g['count'], 'pct': g['pct']} for g in result.get('gt', [])}
            st.markdown('#### 📐 コードブック')
            render_codebook_structure(result['codebook'], gt_by_code=gt_by_code_current, key='codebook_current')
            if result.get('pending_edit'):
                st.caption('📝 未確定の編集案があります。「🔧 コードブック編集」タブで確認してください。')
            st.divider()

        # ── コーディング結果（1件以上コーディング済みの場合のみ表示） ──────
        if coded_count > 0:
            st.subheader('📊 コーディング結果')

            sent               = result['sent']
            unassigned         = result.get('unassigned', 0)
            risk_counts        = result.get('risk_counts', {})
            result_risks       = result.get('enabled_risks', [])
            answer_type_counts = result.get('answer_type_counts', {})
            def _pct(cnt):
                return f'{cnt/coded_count*100:.1f}%' if coded_count else '0.0%'

            answer_type_icons = {'unclear': '❓', 'unanswered': '⬜'}
            answer_type_html = ''.join(
                f"  <div>{answer_type_icons.get(o['key'], '❔')} <b>{o['label']}</b>　"
                f"{answer_type_counts.get(o['key'], 0)}件（{_pct(answer_type_counts.get(o['key'], 0))}）</div>\n"
                for o in ANSWER_TYPE_OPTIONS
            )
            risk_icons = {'claim': '📣', 'personal': '🪪', 'org': '🏢', 'danger': '🚨'}
            risk_html = ''.join(
                f"  <div>{risk_icons.get(o['key'], '⚠️')} <b>{o['label']}</b>　"
                f"{risk_counts.get(o['key'], 0)}件（{_pct(risk_counts.get(o['key'], 0))}）</div>\n"
                for o in RISK_CHECK_OPTIONS if o['key'] in result_risks
            )
            with st.expander('📌 特記情報', expanded=True):
                st.markdown(f"""
    <div style="display:flex; flex-wrap:nowrap; overflow-x:auto; gap:32px; padding:4px 0;">
      <div>😊 <b>ポジティブ</b>　{sent['positive']}件（{_pct(sent['positive'])}）</div>
      <div>😞 <b>ネガティブ</b>　{sent['negative']}件（{_pct(sent['negative'])}）</div>
      <div>😐 <b>ニュートラル</b>　{sent['neutral']}件（{_pct(sent['neutral'])}）</div>
      <div>🌗 <b>混在</b>　{sent.get('mixed', 0)}件（{_pct(sent.get('mixed', 0))}）</div>
      <div>➖ <b>非該当（コードなし）</b>　{unassigned}件（{_pct(unassigned)}）</div>
    {answer_type_html}{risk_html}</div>
    """, unsafe_allow_html=True)

            st.divider()

            gt = result['gt']
            import pandas as pd

            with st.expander('カテゴリ別集計を表示'):
                cat_summary = {}
                for item in gt:
                    cid = item['cat_id']
                    if cid not in cat_summary:
                        cat_summary[cid] = {'カテゴリID': cid, 'カテゴリ名': item['cat_name'], '件数': 0}
                    cat_summary[cid]['件数'] += item['count']
                df_cat = pd.DataFrame(list(cat_summary.values()))
                df_cat['出現率(%)'] = (df_cat['件数'] / coded_count * 100).round(1)
                df_cat = df_cat.sort_values('件数', ascending=False).reset_index(drop=True)
                st.dataframe(df_cat, width='stretch', hide_index=True)

            st.divider()

            st.markdown('#### 📈 コード別GT集計')

            sort_mode = st.radio(
                '表示順',
                ['順A：カテゴリ出現率順→コード出現率順', '順B：コード出現率が多い順'],
                horizontal=True
            )

            _cat_order, _cat_color_full = _category_color_map(gt, result['codes'])
            # カテゴリの並び順は、生の出現数合計（cat_total）ではなく、_category_color_mapが
            # 確定させた順位（cat_rank）で決める。2つのカテゴリの合計が同数タイになった場合、
            # 生の合計値だけをソートキーにすると「タイの間はコード件数だけで全体を横断比較」して
            # しまい、同じカテゴリのコード同士が分断されて隣り合わなくなる（色分けが乱れて見える
            # 不具合の原因）。ランク（整数の順位）を使えばタイでも必ずカテゴリごとに固まる。
            _cat_rank = {cid: i for i, cid in enumerate(_cat_order)}

            if sort_mode == '順B：コード出現率が多い順':
                gt_sorted = sorted(gt, key=lambda x: x['count'], reverse=True)
            else:
                gt_sorted = sorted(gt, key=lambda x: (_cat_rank.get(x['cat_id'], len(_cat_order)), -x['count']))

            import plotly.express as px
            df_plot = pd.DataFrame(gt_sorted)[['cat_name','code_name','count','pct']]
            df_plot.columns = ['カテゴリ','コード名','件数','出現率(%)']
            # 色マップのキーは、df_plot（gt由来）と必ず同じcat_nameソースから作る。
            # result['codes']（現在のコードブック）から作ると、コードブック編集でカテゴリ名を
            # 変更した後（gt側は編集前の名前のまま）に名前が食い違い、Plotlyが該当カテゴリの色を
            # 見つけられず自動配色にフォールバックしてしまう（「グラフの色がたまに乱れる」不具合の原因）。
            _cat_name_map = {g['cat_id']: g['cat_name'] for g in gt}
            color_discrete_map = {_cat_name_map[cid]: f'#{_cat_color_full[cid]}' for cid in _cat_order if cid in _cat_name_map}
            fig = px.bar(
                df_plot,
                x='コード名',
                y='出現率(%)',
                color='カテゴリ',
                category_orders={'コード名': df_plot['コード名'].tolist()},
                color_discrete_map=color_discrete_map,
                labels={'出現率(%)': '出現率(%)', 'コード名': ''},
            )
            fig.update_layout(
                xaxis_tickangle=-45,
                legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                height=500,
                margin=dict(b=120),
            )
            st.plotly_chart(fig, width='stretch')

            st.divider()

            partial_note = '' if coded_count >= total_items else f'（{coded_count}/{total_items}件分）'
            st.markdown(f'#### 💾 レポートのダウンロード{partial_note}')
            excel_bytes = create_excel(
                q_name, gt, sent, coded_count,
                result['results'], result['items'][:coded_count], result['codes'], unassigned,
                risk_counts=result.get('risk_counts', {}), enabled_risks=result.get('enabled_risks', []),
                answer_type_counts=result.get('answer_type_counts', {}),
            )
            st.download_button(
                label='📥 Excelレポートをダウンロード',
                data=excel_bytes,
                file_name=f'AfterCoding_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                width='stretch',
            )

with tab_basic:
    # tab_basicはtab_homeの外（タブバー自体は常時表示のため）。プロジェクトが無い状態でも
    # このタブ自体は選択できるが、中身は「まだ無い」旨の案内のみ表示する。
    if active_result:
        _render_basic_table_tab(result)
    else:
        st.info('コーディング結果がまだありません。「🏠 ホーム」タブの作業メニューでコードブックの作成・コーディングを行ってください。')

with tab_edit:
    if active_result:
        _render_codebook_edit_tab(result, api_key, q_name, diagnostic_size)
    else:
        st.info('コードブックがまだありません。「🏠 ホーム」タブの作業メニューでコードブックの作成・コーディングを行ってください。')

with tab_cross:
    if active_result:
        _render_cross_tab_tab(result)
    else:
        st.info('コーディング結果がまだありません。「🏠 ホーム」タブの作業メニューでコードブックの作成・コーディングを行ってください。')

with tab_bubble:
    if active_result:
        _render_bubble_tab(result)
    else:
        st.info('コーディング結果がまだありません。「🏠 ホーム」タブの作業メニューでコードブックの作成・コーディングを行ってください。')
